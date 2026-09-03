"""Reporte por vuelo (2-sep-2026): el combustible del AVIÓN se rotula
'Gasavión / Turbosina' (ya no 'Gasolina', que ahora nombra la categoría de
vehículos) y la columna Categoría de los gastos imprime la etiqueta amable
cuando el API la manda, o el código del enum si no. El PDF se valida por su
HTML (WeasyPrint no hace falta)."""

from io import BytesIO

from openpyxl import load_workbook

from app.schemas.reportes import ReporteVueloRequest
from app.services.reporte_vuelo_pdf import _build_html
from app.services.reporte_vuelo_xlsx import render_reporte_vuelo_xlsx


def _req() -> ReporteVueloRequest:
    return ReporteVueloRequest(
        generado="2026-09-02",
        folio="131",
        cliente="Cliente Demo",
        total_usd=1160.0,
        iva_usd=160.0,
        tc_usd_mxn=18.0,
        combustible=[
            {"fecha": "2026-08-20", "detalle": "CUN · 100 L", "moneda": "MXN",
             "monto": 3000.0, "litros": 100.0},
        ],
        combustible_total_usd=166.67,
        gastos=[
            {"fecha": "2026-08-20", "concepto": "OTRO",
             "etiqueta": "Otros gastos VuelaTour", "moneda": "MXN", "monto": 500.0},
            # Sin etiqueta (API viejo): se imprime el código.
            {"fecha": "2026-08-20", "concepto": "TUAS", "moneda": "MXN", "monto": 800.0},
        ],
        gastos_total_usd=72.22,
        remanente_usd=761.11,
        ganancia_final_usd=761.11,
    )


def _valores(xlsx: bytes) -> list[str]:
    ws = load_workbook(BytesIO(xlsx)).active
    return [str(c.value) for fila in ws.iter_rows() for c in fila if c.value is not None]


def test_xlsx_rotula_gasavion_turbosina_y_etiquetas() -> None:
    valores = _valores(render_reporte_vuelo_xlsx(_req()))
    assert "Gasavión / Turbosina" in valores
    assert "Total gasavión / turbosina USD" in valores
    assert "(−) Gasavión / Turbosina" in valores
    assert not any("gasolina" in v.lower() for v in valores)
    assert "Otros gastos VuelaTour" in valores
    assert "OTRO" not in valores
    assert "TUAS" in valores


def test_pdf_html_rotula_gasavion_turbosina_y_etiquetas() -> None:
    html = _build_html(_req())
    assert "Total gasavión / turbosina:" in html
    assert "(−) Gasavión / Turbosina" in html
    assert "gasolina" not in html.lower()
    assert "Otros gastos VuelaTour" in html
    assert "<td>OTRO</td>" not in html
    assert "<td>TUAS</td>" in html
