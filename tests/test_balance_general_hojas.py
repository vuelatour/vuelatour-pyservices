"""Hojas de gastos de los dos libros de balance.

GENERAL (renombre 1-sep-2026, modelo mental del equipo): el payload
`gastos_empresa` (gastos sin avión ni vuelo) se pinta en la hoja 'otros
gastos' (antes 'gastos VuelaTour') y los parciales del reparto manual por
avión en 'repartidos a aviones' (antes 'otros gastos'). Ese libro CONSERVA
ambas hojas.

INDIVIDUAL (pedido del cliente 2-sep-2026: 'gastos indirectos' y 'otros
gastos' se confundían): UNA sola pestaña "Gastos Indirectos" con las dos
listas del API fusionadas (presentación pura: el contrato `gastos_indirectos`
/ `otros_gastos` y la cascada no cambian) y una sola fila "(−) GASTOS
INDIRECTOS USD" en la hoja balance = indirectos + otros, con nota cuando hay
reparto manual."""

from io import BytesIO

from openpyxl import load_workbook

from app.schemas.reportes import (
    BalanceAvionHojaGastos,
    BalanceAvionRequest,
    BalanceGeneralRequest,
)
from app.services.balance_avion_xlsx import (
    LIGHT,
    _fusionar_hojas_gastos,
    _suma_none,
    render_balance_avion_xlsx,
    render_balance_general_xlsx,
)

_GASTOS_EMPRESA = {
    "filas": [
        {
            "fecha": "2026-08-05",
            "categoria": "INDIRECTO",
            "detalle": "Renta de oficina agosto",
            "monto_mxn": 12000.0,
        },
    ],
    "total_mxn": 12000.0,
    "usd": 600.0,
    "usd_hr": None,
}

_OTROS_GASTOS = {
    "filas": [
        {
            "fecha": "2026-08-15",
            "categoria": "INDIRECTO",
            "detalle": "XB-ABC · Nómina repartida (50 %)",
            "monto_mxn": 5000.0,
            "matricula": "XB-ABC",
            "avion_color": "#FF0000",
        },
    ],
    "total_mxn": 5000.0,
    "usd": 250.0,
    "usd_hr": None,
}

# Libro INDIVIDUAL: la lista `gastos_indirectos` (gastos con avión sin vuelo)
# y la lista `otros_gastos` (parciales del reparto manual, con la marca que
# deja el API en el DETALLE) — llegan APARTE y se pintan juntas.
_INDIRECTOS_INDIVIDUAL = {
    "filas": [
        {
            "fecha": "2026-08-20",
            "categoria": "Mantenimiento",
            "detalle": "Lavado de avión",
            "monto_mxn": 800.0,
        },
        {
            "fecha": "2026-08-02",
            "categoria": "Servicios",
            "detalle": "Hangaraje agosto",
            "monto_mxn": 1200.0,
        },
    ],
    "total_mxn": 2000.0,
    "usd": 100.0,
    "usd_hr": 10.0,
}
_OTROS_INDIVIDUAL = {
    "filas": [
        {
            "fecha": "2026-08-15",
            "categoria": "Nómina",
            "detalle": "Nómina agosto · reparto manual: $5,000.00 de $10,000.00 MXN",
            "monto_mxn": 5000.0,
        },
    ],
    "total_mxn": 5000.0,
    "usd": 250.0,
    "usd_hr": None,
}
_BALANCE_INDIVIDUAL = {
    "utilidad_antes_usd": 1000.0,
    "combustible_usd": 80.0,
    "gastos_indirectos_usd": 100.0,
    "refacciones_usd": 20.0,
    "otros_usd": 250.0,
    "permisos_usd": 50.0,
    # Cascada YA calculada por el API (resta indirectos y otros UNA vez).
    "utilidad_despues_usd": 500.0,
    "por_cobrar_usd": 0.0,
    "utilidad_cobrada_usd": 500.0,
}


def _general(**extra) -> BalanceGeneralRequest:
    return BalanceGeneralRequest(
        periodo_desde="2026-08-01",
        periodo_hasta="2026-08-31",
        consolidado=BalanceAvionRequest(
            matricula="FLOTA",
            periodo_desde="2026-08-01",
            periodo_hasta="2026-08-31",
            otros_gastos=_OTROS_GASTOS,
        ),
        **extra,
    )


def _individual(**extra) -> BalanceAvionRequest:
    return BalanceAvionRequest(
        matricula="XB-ABC",
        periodo_desde="2026-08-01",
        periodo_hasta="2026-08-31",
        gastos_indirectos=_INDIRECTOS_INDIVIDUAL,
        otros_gastos=_OTROS_INDIVIDUAL,
        balance=_BALANCE_INDIVIDUAL,
        **extra,
    )


def _filas_gastos(ws) -> list[tuple]:
    """(fecha, categoría, detalle, monto) de las filas del ledger (desde la
    fila 8, hasta la primera vacía)."""
    filas = []
    row = 8
    while ws.cell(row=row, column=1).value is not None:
        filas.append(tuple(ws.cell(row=row, column=c).value for c in range(1, 5)))
        row += 1
    return filas


def _bloque_balance(ws, desde: int = 4) -> dict[str, tuple]:
    """etiqueta → (monto, texto de la nota o None) de las filas de la
    cascada, hasta el título del reparto a socios."""
    bloque = {}
    row = desde
    while True:
        label = ws.cell(row=row, column=1).value
        if label is None or str(label).startswith("Reparto a socios"):
            break
        cell = ws.cell(row=row, column=2)
        bloque[label] = (cell.value, cell.comment.text if cell.comment else None)
        row += 1
    return bloque


# ===== GENERAL: conserva sus dos hojas =====


def test_general_renombra_hojas_de_gastos():
    data = render_balance_general_xlsx(_general(gastos_empresa=_GASTOS_EMPRESA))
    wb = load_workbook(BytesIO(data))

    # Nombres nuevos (1-sep-2026); el viejo 'gastos VuelaTour' ya no existe.
    assert "otros gastos" in wb.sheetnames
    assert "repartidos a aviones" in wb.sheetnames
    assert "gastos VuelaTour" not in wb.sheetnames
    # Pestañas hermanas: 'otros gastos' (empresa) va JUNTO y ANTES de
    # 'repartidos a aviones' (la posición que el equipo ya conocía).
    idx = wb.sheetnames.index
    assert idx("otros gastos") + 1 == idx("repartidos a aviones")

    # 'otros gastos' = los gastos de EMPRESA (payload gastos_empresa), con el
    # título exacto (sin sufijo de matrícula) y su fila.
    ws = wb["otros gastos"]
    assert ws.cell(row=1, column=1).value == (
        "Otros gastos — VuelaTour (sin avión ni vuelo)"
    )
    detalles = {c[2].value for c in ws.iter_rows(min_col=1, max_col=4)}
    assert "Renta de oficina agosto" in detalles

    # 'repartidos a aviones' = los parciales del reparto manual por avión
    # (payload otros_gastos del consolidado, que NO cambia de nombre).
    ws_r = wb["repartidos a aviones"]
    assert ws_r.cell(row=1, column=1).value == (
        "Otros gastos repartidos a aviones — FLOTA"
    )
    detalles_r = {c[2].value for c in ws_r.iter_rows(min_col=1, max_col=4)}
    assert "XB-ABC · Nómina repartida (50 %)" in detalles_r


def test_general_sin_gastos_empresa_conserva_repartidos():
    # API viejo (skew): sin `gastos_empresa` no hay hoja de empresa, pero los
    # parciales se pintan igual con el nombre nuevo.
    data = render_balance_general_xlsx(_general())
    wb = load_workbook(BytesIO(data))
    assert "otros gastos" not in wb.sheetnames
    assert "gastos VuelaTour" not in wb.sheetnames
    assert "repartidos a aviones" in wb.sheetnames


def test_general_no_fusiona_y_titula_balance_general_vuelatour():
    # La fusión del 2-sep es SOLO del libro individual: el general conserva
    # 'otros gastos' + 'repartidos a aviones' y no crea "Gastos Indirectos".
    data = render_balance_general_xlsx(_general(gastos_empresa=_GASTOS_EMPRESA))
    wb = load_workbook(BytesIO(data))
    assert "Gastos Indirectos" not in wb.sheetnames
    assert "otros gastos" in wb.sheetnames
    assert "repartidos a aviones" in wb.sheetnames
    # Título A1 del RESUMEN = nombre del apartado nuevo en Reportes.
    assert wb["RESUMEN flota"].cell(row=1, column=1).value == (
        "Balance general VuelaTour · 2026-08-01 a 2026-08-31"
    )
    assert wb["balance"].cell(row=1, column=1).value == (
        "Balance por avión — Balance general VuelaTour"
    )
    # En 'repartidos a aviones' la celda DETALLE lleva el color del AVIÓN,
    # no el tinte de parciales (ese es solo del libro individual).
    ws_r = wb["repartidos a aviones"]
    assert ws_r.cell(row=8, column=3).fill.fgColor.rgb == "00FF0000"


def test_general_bloque_balance_una_sola_fila_de_indirectos():
    data = render_balance_general_xlsx(
        _general(
            aviones=[
                BalanceAvionRequest(
                    matricula="XB-ABC",
                    balance={"gastos_indirectos_usd": 100.0, "otros_usd": 50.0},
                ),
                BalanceAvionRequest(
                    matricula="XB-DEF",
                    balance={"gastos_indirectos_usd": 30.0, "otros_usd": 0.0},
                ),
            ]
        )
    )
    ws = load_workbook(BytesIO(data))["balance"]
    etiquetas = [
        ws.cell(row=r, column=1).value
        for r in range(1, ws.max_row + 1)
        if ws.cell(row=r, column=1).value
    ]
    assert etiquetas.count("(−) GASTOS INDIRECTOS USD") == 2
    assert "(−) OTROS GASTOS REPARTIDOS AL AVIÓN USD" not in etiquetas
    assert "(−) OTROS GASTOS USD" not in etiquetas
    # Bloque XB-ABC (título en la fila 4, cascada desde la 5): 100 + 50 con
    # nota que apunta a 'repartidos a aviones'.
    b1 = _bloque_balance(ws, desde=5)
    monto, nota = b1["(−) GASTOS INDIRECTOS USD"]
    assert monto == 150.0
    assert nota is not None
    assert "$50.00 USD" in nota and "repartidos a aviones" in nota
    # Bloque XB-DEF: otros_usd = 0 → sin nota.
    fila_def = next(
        r for r in range(1, ws.max_row + 1) if ws.cell(row=r, column=1).value == "XB-DEF"
    )
    b2 = _bloque_balance(ws, desde=fila_def + 1)
    assert b2["(−) GASTOS INDIRECTOS USD"] == (30.0, None)


# ===== INDIVIDUAL: pestaña única "Gastos Indirectos" =====


def test_suma_none_tolerante():
    assert _suma_none(None, None) is None
    assert _suma_none(None, 5.0) == 5.0
    assert _suma_none(5.0, None) == 5.0
    assert _suma_none(1.1, 2.2) == 3.3
    assert _suma_none(0.0, 0.0) == 0.0


def test_fusionar_hojas_gastos_ordena_por_fecha_y_suma_totales():
    a = BalanceAvionHojaGastos(**_INDIRECTOS_INDIVIDUAL)
    b = BalanceAvionHojaGastos(**_OTROS_INDIVIDUAL)
    fusion = _fusionar_hojas_gastos(a, b)
    assert [f.detalle for f in fusion.filas] == [
        "Hangaraje agosto",
        "Nómina agosto · reparto manual: $5,000.00 de $10,000.00 MXN",
        "Lavado de avión",
    ]
    assert fusion.total_mxn == 7000.0
    # usd y usd_hr son la suma de lo que YA mandó el API (None + x = x) —
    # jamás se recalcula el USD desde el MXN.
    assert fusion.usd == 350.0
    assert fusion.usd_hr == 10.0
    # Las hojas de entrada no se tocan y el sort es estable (misma fecha →
    # primero las de `a`, en su orden).
    assert [f.detalle for f in a.filas] == ["Lavado de avión", "Hangaraje agosto"]
    sin_fecha = _fusionar_hojas_gastos(
        BalanceAvionHojaGastos(filas=[{"detalle": "a1"}, {"detalle": "a2"}]),
        BalanceAvionHojaGastos(filas=[{"detalle": "b1"}]),
    )
    assert [f.detalle for f in sin_fecha.filas] == ["a1", "a2", "b1"]
    assert sin_fecha.total_mxn is None and sin_fecha.usd is None


def test_individual_fusiona_indirectos_y_otros_en_una_pestana():
    data = render_balance_avion_xlsx(_individual())
    wb = load_workbook(BytesIO(data))

    # Pestaña exacta "Gastos Indirectos"; ya no existen 'gastos indirectos'
    # ni 'otros gastos' (ni la del general).
    assert "Gastos Indirectos" in wb.sheetnames
    assert "gastos indirectos" not in wb.sheetnames
    assert "otros gastos" not in wb.sheetnames
    assert "repartidos a aviones" not in wb.sheetnames
    assert wb.sheetnames == [
        "reporte horas XB-ABC", "cobranza", "combustible", "Gastos Indirectos",
        "permisos", "balance", "pendientes de captura",
    ]

    ws = wb["Gastos Indirectos"]
    assert ws.cell(row=1, column=1).value == "Gastos indirectos — XB-ABC"
    # Totales = suma de los dos totales del API (MXN, USD y USD/hr).
    assert ws.cell(row=5, column=1).value == 7000.0
    assert ws.cell(row=5, column=3).value == 350.0
    assert ws.cell(row=5, column=5).value == 10.0
    # Filas de AMBAS listas, ordenadas por fecha.
    assert _filas_gastos(ws) == [
        ("02/08/2026", "Servicios", "Hangaraje agosto", 1200.0),
        (
            "15/08/2026", "Nómina",
            "Nómina agosto · reparto manual: $5,000.00 de $10,000.00 MXN", 5000.0,
        ),
        ("20/08/2026", "Mantenimiento", "Lavado de avión", 800.0),
    ]

    # Hoja balance: UNA fila de indirectos = 100 + 250, con la nota del
    # reparto manual; sin fila de OTROS GASTOS. El resto de la cascada
    # intacto (viene tal cual del API).
    bloque = _bloque_balance(wb["balance"])
    assert "(−) OTROS GASTOS USD" not in bloque
    assert "(−) OTROS GASTOS REPARTIDOS AL AVIÓN USD" not in bloque
    monto, nota = bloque["(−) GASTOS INDIRECTOS USD"]
    assert monto == 350.0
    assert nota == (
        "Incluye $250.00 USD de gastos administrativos repartidos a mano "
        "(filas reparto manual de la hoja Gastos Indirectos)."
    )
    assert bloque["(−) COMBUSTIBLE DEL MES USD"] == (80.0, None)
    assert bloque["(−) REFACCIONES (INVENTARIO) USD"] == (20.0, None)
    assert bloque["(−) PERMISOS USD"] == (50.0, None)
    assert bloque["UTILIDAD DESPUÉS DE GASTOS USD"] == (500.0, None)
    assert bloque["UTILIDAD COBRADA USD"] == (500.0, None)


def test_individual_refacciones_va_entre_gastos_indirectos_y_permisos():
    data = render_balance_avion_xlsx(
        _individual(refacciones={"filas": [], "total_mxn": None, "usd": None})
    )
    wb = load_workbook(BytesIO(data))
    assert wb.sheetnames == [
        "reporte horas XB-ABC", "cobranza", "combustible", "Gastos Indirectos",
        "refacciones", "permisos", "balance", "pendientes de captura",
    ]


def test_individual_sin_otros_no_lleva_nota_en_balance():
    # otros_usd None (API viejo) o 0: la fila trae solo los indirectos y la
    # celda no lleva nota; None + None = celda vacía (nunca un 0 falso).
    for otros in (None, 0.0):
        data = render_balance_avion_xlsx(
            BalanceAvionRequest(
                matricula="XB-ABC",
                balance={"gastos_indirectos_usd": 100.0, "otros_usd": otros},
            )
        )
        bloque = _bloque_balance(load_workbook(BytesIO(data))["balance"])
        assert bloque["(−) GASTOS INDIRECTOS USD"] == (100.0, None)
    data = render_balance_avion_xlsx(BalanceAvionRequest(matricula="XB-ABC"))
    bloque = _bloque_balance(load_workbook(BytesIO(data))["balance"])
    assert bloque["(−) GASTOS INDIRECTOS USD"] == (None, None)


def test_individual_resalta_solo_las_filas_de_reparto_manual():
    data = render_balance_avion_xlsx(_individual())
    ws = load_workbook(BytesIO(data))["Gastos Indirectos"]
    fills = {
        ws.cell(row=r, column=3).value: ws.cell(row=r, column=3).fill
        for r in range(8, 11)
    }
    parcial = fills["Nómina agosto · reparto manual: $5,000.00 de $10,000.00 MXN"]
    assert parcial.fill_type == "solid"
    assert parcial.fgColor.rgb == f"00{LIGHT}"
    # Los gastos capturados directo al avión NO llevan tinte.
    assert fills["Hangaraje agosto"].fill_type is None
    assert fills["Lavado de avión"].fill_type is None
    # Sin columnas nuevas: el ancho del DETALLE sigue en la columna C.
    assert ws.column_dimensions["C"].width == 70
