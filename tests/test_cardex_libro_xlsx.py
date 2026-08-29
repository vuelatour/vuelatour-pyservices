"""Tests del cardex formato LIBRO: encabezado doble, filas por lado y totales."""

from io import BytesIO

from openpyxl import load_workbook

from app.schemas.cardex_libro import CardexLibroRequest
from app.services.cardex_libro_xlsx import render_cardex_libro_xlsx


def _req() -> CardexLibroRequest:
    return CardexLibroRequest(
        titulo="Cardex — Filtro 108-1",
        item_nombre="Filtro 108-1",
        numero_parte="108-1",
        unidad="pieza",
        generado="2026-08-29",
        moneda="MXN",
        entradas=[
            {
                "fecha": "2026-08-01",
                "cantidad": 10,
                "descripcion": "Filtro 108-1 · Aircraft Spruce",
                "valor_compra_unitario": 500.0,
                "valor_compra_total": 5000.0,
                "stock_despues": 10,
            },
        ],
        salidas=[
            {
                "fecha": "2026-08-10",
                "cantidad": 2,
                "descripcion": "Filtro 108-1",
                "venta_unitaria": 800.0,
                "venta_total": 1600.0,
                "remanente": 8,
                "ganancia": 600.0,
                "vendido_a": "XB-ABC",
            },
            {
                "fecha": "2026-08-12",
                "cantidad": 1,
                "descripcion": "Filtro 108-1 · a costo FIFO",
                "venta_unitaria": 500.0,
                "venta_total": 500.0,
                "remanente": 7,
                "ganancia": 0.0,
                "vendido_a": "FLOTA",
            },
        ],
        total_compra=5000.0,
        total_venta=2100.0,
        total_ganancia=600.0,
    )


def _hoja(xlsx: bytes):
    return load_workbook(BytesIO(xlsx)).active


def test_encabezado_doble_y_columnas() -> None:
    ws = _hoja(render_cardex_libro_xlsx(_req()))
    # Fila 4: bloques; fila 5: columnas de cada lado.
    assert ws.cell(row=4, column=1).value == "ENTRADAS"
    assert ws.cell(row=4, column=7).value == "SALIDAS"
    assert ws.cell(row=5, column=1).value == "Fecha entrada"
    assert ws.cell(row=5, column=6).value == "Cantidad en stock"
    assert ws.cell(row=5, column=7).value == "Fecha de salida"
    assert ws.cell(row=5, column=10).value == "Valor unitario al que se vendió"
    assert ws.cell(row=5, column=13).value == "Ganancia"
    assert ws.cell(row=5, column=14).value == "A quién se le vendió"


def test_filas_por_lado_y_totales() -> None:
    ws = _hoja(render_cardex_libro_xlsx(_req()))
    # Fila 6: entrada 1 a la izquierda, salida 1 a la derecha.
    assert ws.cell(row=6, column=1).value == "2026-08-01"
    assert ws.cell(row=6, column=5).value == 5000.0
    assert ws.cell(row=6, column=6).value == 10
    assert ws.cell(row=6, column=7).value == "2026-08-10"
    assert ws.cell(row=6, column=10).value == 800.0
    assert ws.cell(row=6, column=13).value == 600.0
    assert ws.cell(row=6, column=14).value == "XB-ABC"
    # Fila 7: sin entrada (lado izquierdo vacío), salida 2 a la derecha.
    assert ws.cell(row=7, column=1).value is None
    assert ws.cell(row=7, column=14).value == "FLOTA"
    # Fila 8: totales (compra | venta | ganancia), los mandó el API.
    assert ws.cell(row=8, column=5).value == 5000.0
    assert ws.cell(row=8, column=11).value == 2100.0
    assert ws.cell(row=8, column=13).value == 600.0
