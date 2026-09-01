"""Hoja 'inventario' (tiendita) del Balance GENERAL: bloque por ítem +
detalle de salidas; sustituye a 'refacciones' SOLO en el general (fallback
skew tolerante con API viejo) y el libro INDIVIDUAL conserva la suya."""

from io import BytesIO

from openpyxl import load_workbook

from app.schemas.reportes import (
    BalanceAvionRequest,
    BalanceGeneralRequest,
)
from app.services.balance_avion_xlsx import (
    render_balance_avion_xlsx,
    render_balance_general_xlsx,
)

_REFACCIONES = {
    "filas": [
        {
            "fecha": "2026-08-10",
            "categoria": "Refacción",
            "detalle": "XB-ABC · Salida de bodega: 2 × Filtro 108-1 (precio de venta)",
            "monto_mxn": 1600.0,
            "moneda_original": None,
            "monto_original": None,
            "matricula": "XB-ABC",
            "avion_color": "#FF0000",
            "costo_mxn": 1000.0,
            "venta_mxn": 1600.0,
        },
        {
            "fecha": "2026-08-12",
            "categoria": "Refacción",
            "detalle": "XB-DEF · Salida de bodega: 1 × Aceite (costo FIFO)",
            "monto_mxn": 500.0,
            "moneda_original": None,
            "monto_original": None,
            "matricula": "XB-DEF",
            "costo_mxn": 500.0,
            "venta_mxn": 500.0,
        },
    ],
    "total_mxn": 2100.0,
    "usd": 110.0,
    "usd_hr": None,
}

_INVENTARIO = {
    "filas": [
        {
            "nombre": "Filtro 108-1 · 108-1",
            "existencia": 8,
            "valor_costo_mxn": 4000.0,
            "compradas_cant": 10,
            "compradas_costo_mxn": 5000.0,
            "salidas_cant": 2,
            "vendido_mxn": 1600.0,
            "utilidad_mxn": 600.0,
            "matriculas": "XB-ABC + FLOTA",
        },
        {
            # Ítem con stock pero sin actividad del periodo: solo existencia.
            "nombre": "Bujía fina",
            "existencia": 4,
            "valor_costo_mxn": 800.0,
        },
    ],
    "total_piezas": 12,
    "total_valor_mxn": 4800.0,
    "total_compras_mxn": 5000.0,
    "total_vendido_mxn": 1600.0,
    "total_utilidad_mxn": 600.0,
}


def _general(**extra) -> BalanceGeneralRequest:
    return BalanceGeneralRequest(
        periodo_desde="2026-08-01",
        periodo_hasta="2026-08-31",
        consolidado=BalanceAvionRequest(
            matricula="FLOTA",
            periodo_desde="2026-08-01",
            periodo_hasta="2026-08-31",
            refacciones=_REFACCIONES,
        ),
        **extra,
    )


def _sheet(data: bytes, nombre: str):
    return load_workbook(BytesIO(data))[nombre]


def test_general_con_inventario_pinta_hoja_y_no_refacciones():
    data = render_balance_general_xlsx(_general(inventario=_INVENTARIO))
    wb = load_workbook(BytesIO(data))
    assert "inventario" in wb.sheetnames
    assert "refacciones" not in wb.sheetnames

    ws = wb["inventario"]
    # Bloque 1: encabezados, fila por ítem y fila TOTALES.
    assert ws.cell(row=4, column=1).value == "ÍTEM"
    assert ws.cell(row=4, column=9).value == "MATRÍCULAS"
    assert ws.cell(row=5, column=1).value == "Filtro 108-1 · 108-1"
    assert ws.cell(row=5, column=2).value == 8
    assert ws.cell(row=5, column=5).value == 5000.0
    assert ws.cell(row=5, column=9).value == "XB-ABC + FLOTA"
    # None = celda vacía (ítem sin actividad del periodo), jamás 0 falso.
    assert ws.cell(row=6, column=4).value is None
    assert ws.cell(row=6, column=7).value is None
    assert ws.cell(row=7, column=1).value == "TOTALES"
    assert ws.cell(row=7, column=2).value == 12
    assert ws.cell(row=7, column=3).value == 4800.0
    assert ws.cell(row=7, column=8).value == 600.0

    # Bloque 2: detalle de salidas (la vieja hoja 'refacciones'): la columna
    # AVIÓN lleva la matrícula y el detalle pierde su prefijo 'MATRÍCULA · '.
    filas = {
        (c[0].value, c[1].value, c[2].value, c[3].value, c[4].value, c[5].value)
        for c in ws.iter_rows(min_col=1, max_col=6)
    }
    assert (
        "10/08/2026",
        "Salida de bodega: 2 × Filtro 108-1 (precio de venta)",
        "XB-ABC",
        1000.0,
        1600.0,
        600.0,  # GANANCIA = venta − costo (solo para mostrar)
    ) in filas
    assert (
        "12/08/2026",
        "Salida de bodega: 1 × Aceite (costo FIFO)",
        "XB-DEF",
        500.0,
        500.0,
        0.0,
    ) in filas
    # TOTALES del bloque 2 (re-suma de columnas, solo para mostrar).
    assert ("TOTALES", None, None, 1500.0, 2100.0, 600.0) in filas


def test_general_sin_inventario_conserva_refacciones():
    # API viejo (skew): sin `inventario` la hoja 'refacciones' sigue igual.
    data = render_balance_general_xlsx(_general())
    wb = load_workbook(BytesIO(data))
    assert "refacciones" in wb.sheetnames
    assert "inventario" not in wb.sheetnames


def test_individual_conserva_su_hoja_refacciones():
    data = render_balance_avion_xlsx(
        BalanceAvionRequest(matricula="XB-ABC", refacciones=_REFACCIONES)
    )
    wb = load_workbook(BytesIO(data))
    assert "refacciones" in wb.sheetnames
    assert "inventario" not in wb.sheetnames
