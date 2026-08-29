"""Tests del export genérico tabla-xlsx: contrato base intacto + resaltes."""

from io import BytesIO

from openpyxl import load_workbook

from app.schemas.tabla import TablaXlsxRequest
from app.services.tabla_xlsx import render_tabla_xlsx


def _req(**extra) -> TablaXlsxRequest:
    return TablaXlsxRequest(
        titulo="Conciliación · Prueba",
        subtitulo="2 movimientos",
        columnas=[
            {"label": "Fecha", "tipo": "texto"},
            {"label": "Cargo", "tipo": "money"},
        ],
        filas=[
            ["2026-08-01", 100.5],
            ["2026-08-02", 200.0],
        ],
        totales=["Totales", 300.5],
        **extra,
    )


def _hoja(xlsx: bytes):
    return load_workbook(BytesIO(xlsx)).active


def test_tabla_sin_resaltes_render_intacto() -> None:
    """Sin `resaltes` el render es el de siempre (contrato de los 4 callers)."""
    ws = _hoja(render_tabla_xlsx(_req()))
    # Título fila 1, subtítulo fila 2, blanco fila 3, encabezado fila 4.
    assert ws.cell(row=4, column=1).value == "Fecha"
    celda = ws.cell(row=5, column=2)
    assert celda.value == 100.5
    assert not celda.font.bold
    # Sin relleno de resalte.
    assert celda.fill.patternType is None


def test_tabla_con_resaltes_pinta_naranja() -> None:
    """La celda marcada queda bold + naranja con fondo suave; el resto igual."""
    ws = _hoja(
        render_tabla_xlsx(
            _req(
                resaltes=[
                    {"fila": 0, "col": 1},
                    # Fuera de rango: se ignora sin tronar el render.
                    {"fila": 99, "col": 1},
                ]
            )
        )
    )
    marcada = ws.cell(row=5, column=2)
    assert marcada.font.bold
    assert str(marcada.font.color.rgb).endswith("ED7D31")
    assert str(marcada.fill.fgColor.rgb).endswith("FFF3E6")
    # Conserva el formato de moneda de su columna.
    assert marcada.number_format == '"$"#,##0.00'
    # La fila NO marcada queda como siempre.
    libre = ws.cell(row=6, column=2)
    assert not libre.font.bold
    assert libre.fill.patternType is None
