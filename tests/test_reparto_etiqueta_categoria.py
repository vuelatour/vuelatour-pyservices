"""Reparto de utilidades: la categoría de gasto se imprime con su etiqueta
amable (homologada con panel/app/API, 2-sep-2026) y, si el API no la manda,
con el código del enum tal cual. El bloque es ADITIVO: sin
`gastos_por_categoria` el PDF y el Excel salen como siempre."""

from io import BytesIO

from openpyxl import load_workbook

from app.schemas.reparto import RepartoPdfRequest
from app.services.reparto_pdf import _filas_gastos_categoria, render_reparto_pdf
from app.services.reparto_xlsx import render_reparto_xlsx

_GASTOS = [
    {
        "categoria": "GAS",
        "etiqueta": "Gasavión / Turbosina",
        "grupo": "DIRECTO",
        "count": 2,
        "usd": 250.0,
    },
    # Sin etiqueta (API viejo): se imprime el código tal cual.
    {"categoria": "OTRO", "grupo": "FIJO", "count": 1, "usd": 50.0},
]


def _req(gastos: list[dict] | None = None) -> RepartoPdfRequest:
    return RepartoPdfRequest(
        periodo_desde="2026-08-01",
        periodo_hasta="2026-08-31",
        generado="2026-09-02",
        aviones=[
            {
                "matricula": "XB-ABC",
                "modelo": "Cessna 206",
                "ingresos_cobrado_usd": 1000.0,
                "gastos_directos_usd": 300.0,
                "saldo_usd": 700.0,
                "reparto": [
                    {"socio_nombre": "Socio", "porcentaje": 100.0, "monto_usd": 700.0}
                ],
                "gastos_por_categoria": gastos or [],
            }
        ],
    )


def _valores(xlsx: bytes) -> set:
    ws = load_workbook(BytesIO(xlsx)).active
    return {c.value for fila in ws.iter_rows() for c in fila if c.value is not None}


def test_schema_aditivo_sin_bloque() -> None:
    assert _req().aviones[0].gastos_por_categoria == []


def test_pdf_filas_usan_etiqueta_o_categoria() -> None:
    filas = _filas_gastos_categoria(_req(_GASTOS).aviones[0].gastos_por_categoria)
    assert filas == [
        ["Gasavión / Turbosina", "Directo", "2", "$250.00"],
        ["OTRO", "Fijo (manual)", "1", "$50.00"],
    ]


def test_pdf_genera_con_y_sin_bloque() -> None:
    for req in (_req(_GASTOS), _req()):
        pdf = render_reparto_pdf(req)
        assert pdf[:4] == b"%PDF"
        assert len(pdf) > 1000


def test_xlsx_imprime_etiqueta_y_cae_al_codigo() -> None:
    valores = _valores(render_reparto_xlsx(_req(_GASTOS)))
    assert "Gastos por categoría" in valores
    assert "Gasavión / Turbosina" in valores
    assert "GAS" not in valores
    assert "OTRO" in valores
    assert "Directo" in valores
    assert "Fijo (manual)" in valores


def test_xlsx_sin_gastos_no_pinta_seccion() -> None:
    valores = _valores(render_reparto_xlsx(_req()))
    assert "Gastos por categoría" not in valores
