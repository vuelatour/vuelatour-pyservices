"""Carga masiva de combustible: plantilla XLSX y parseo estructural.

La oficina descarga la plantilla (una fila = una carga/ticket), la llena y la
sube. Aquí SOLO se renderiza el libro y se convierte de vuelta a JSON con
tipos básicos; toda la validación de negocio (matrículas reales, folios,
monedas, tipo de cambio) vive en vuelatour-api.
"""

from __future__ import annotations

import base64
import binascii
import csv
import io
import unicodedata
from datetime import date, datetime, time

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from app.schemas.gastos import (
    FilaCombustible,
    ParseCombustibleRequest,
    ParseCombustibleResponse,
    PlantillaCombustibleRequest,
)

HOJA_CARGAS = "Cargas de combustible"
HOJA_CATALOGOS = "Catálogos"
COMPROBANTES = ["FACTURA", "TICKET", "PENDIENTE"]
# Las validaciones de datos cubren las filas 2..500 de la hoja de cargas.
MAX_FILA_DV = 500

BRAND = "0F4C81"
GRIS_EJEMPLO = "9AA0A6"
WHITE = "FFFFFF"

# Contrato de columnas: ORDEN EXACTO de la fila 1 (el API y el panel dependen
# de estos encabezados; el parser los busca por nombre, no por posición).
ENCABEZADOS = [
    "MATRÍCULA*",
    "FECHA DE CARGA* (dd/mm/aaaa)",
    "HORA (HH:MM)",
    "LITROS*",
    "MONTO TOTAL*",
    "MONEDA*",
    "TIPO DE CAMBIO (MXN por USD)",
    "TIPO COMBUSTIBLE",
    "LUGAR",
    "PROVEEDOR",
    "MEDIO DE PAGO*",
    "FOLIO VUELO",
    "COMPROBANTE",
    "NOTAS",
]
ANCHOS = [14, 24, 12, 10, 14, 11, 26, 18, 14, 24, 16, 13, 15, 34]

INSTRUCCIONES = [
    "Una fila = una carga/ticket de combustible.",
    "Los campos con * son obligatorios.",
    "MONTO TOTAL = lo que llega al estado de cuenta (total pagado).",
    "Fechas y horas en hora de Cancún.",
    "Si MONEDA es MXN y no capturas TIPO DE CAMBIO (MXN por USD), la carga "
    "queda fuera del balance en USD hasta que lo captures en el panel.",
    "FOLIO VUELO liga la carga a un vuelo (opcional; déjalo vacío si no aplica).",
    "COMPROBANTE: FACTURA, TICKET o PENDIENTE.",
    "Borra las filas de ejemplo (en gris) antes de subir el archivo.",
    "No cambies los encabezados ni el orden de las columnas.",
]


# --------------------------------------------------------------------------
# Plantilla
# --------------------------------------------------------------------------


def _primero(lista: list[str], default: str) -> str:
    return lista[0] if lista else default


def _segundo(lista: list[str], default: str) -> str:
    return lista[1] if len(lista) > 1 else _primero(lista, default)


def render_plantilla_combustible(req: PlantillaCombustibleRequest) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = HOJA_CARGAS

    # Encabezado (fila 1).
    for col, (titulo, ancho) in enumerate(zip(ENCABEZADOS, ANCHOS, strict=True), start=1):
        cell = ws.cell(row=1, column=col, value=titulo)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=BRAND)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = ancho
    ws.freeze_panes = "A2"

    # Filas 2 y 3: ejemplos realistas (gris itálica) que el usuario debe borrar.
    moneda_mxn = "MXN" if not req.monedas or "MXN" in req.monedas else req.monedas[0]
    moneda_usd = "USD" if not req.monedas or "USD" in req.monedas else moneda_mxn
    ejemplos = [
        [
            _primero(req.matriculas, "XB-ABC"),
            "05/07/2026",
            "09:30",
            180,
            5850.00,
            moneda_mxn,
            17.25,
            _primero(req.tipos_combustible, "AVGAS 100LL"),
            "CUN",
            _primero(req.proveedores, "ASA"),
            _primero(req.medios_pago, "EFECTIVO"),
            1234,
            "FACTURA",
            "(borra estas filas de ejemplo)",
        ],
        [
            _segundo(req.matriculas, "XB-DEF"),
            "06/07/2026",
            "14:15",
            210.5,
            320.50,
            moneda_usd,
            None,
            _segundo(req.tipos_combustible, "JET A-1"),
            "CZM",
            _segundo(req.proveedores, "ASA"),
            _segundo(req.medios_pago, "EFECTIVO"),
            None,
            "TICKET",
            "(borra estas filas de ejemplo)",
        ],
    ]
    fuente_ejemplo = Font(italic=True, color=GRIS_EJEMPLO)
    for r, fila in enumerate(ejemplos, start=2):
        for col, valor in enumerate(fila, start=1):
            cell = ws.cell(row=r, column=col, value=valor)
            cell.font = fuente_ejemplo
            if col in (4, 5, 7):  # LITROS, MONTO TOTAL, TIPO DE CAMBIO
                cell.number_format = "#,##0.00"

    # Hoja de catálogos: de aquí salen los dropdowns (rangos absolutos).
    cat = wb.create_sheet(HOJA_CATALOGOS)
    catalogos = [
        ("MATRÍCULAS", req.matriculas),
        ("MONEDAS", req.monedas),
        ("TIPOS COMBUSTIBLE", req.tipos_combustible),
        ("MEDIOS DE PAGO", req.medios_pago),
        ("COMPROBANTES", COMPROBANTES),
        ("PROVEEDORES", req.proveedores),
    ]
    for col, (titulo, valores) in enumerate(catalogos, start=1):
        head = cat.cell(row=1, column=col, value=titulo)
        head.font = Font(bold=True, color=WHITE)
        head.fill = PatternFill("solid", fgColor=BRAND)
        ancho = max([len(titulo)] + [len(v) for v in valores]) + 4
        cat.column_dimensions[get_column_letter(col)].width = ancho
        for r, valor in enumerate(valores, start=2):
            cat.cell(row=r, column=col, value=valor)
    cat.freeze_panes = "A2"

    # Validaciones de datos (dropdowns) sobre las filas 2..MAX_FILA_DV.
    # columna en "Cargas" -> (columna del catálogo, cantidad de valores)
    dropdowns = [
        ("A", 1, len(req.matriculas)),  # MATRÍCULA
        ("F", 2, len(req.monedas)),  # MONEDA
        ("H", 3, len(req.tipos_combustible)),  # TIPO COMBUSTIBLE
        ("K", 4, len(req.medios_pago)),  # MEDIO DE PAGO
        ("M", 5, len(COMPROBANTES)),  # COMPROBANTE
    ]
    for col_carga, col_cat, n in dropdowns:
        if n == 0:
            continue  # sin catálogo no hay dropdown (campo queda libre)
        letra_cat = get_column_letter(col_cat)
        dv = DataValidation(
            type="list",
            formula1=f"'{HOJA_CATALOGOS}'!${letra_cat}$2:${letra_cat}${n + 1}",
            allow_blank=True,
            showErrorMessage=True,
        )
        dv.errorTitle = "Valor inválido"
        dv.error = "Elige un valor de la lista (hoja Catálogos)."
        ws.add_data_validation(dv)
        dv.add(f"{col_carga}2:{col_carga}{MAX_FILA_DV}")

    # Hoja de instrucciones.
    ins = wb.create_sheet("Instrucciones")
    titulo = ins.cell(row=1, column=1, value="INSTRUCCIONES — CARGA MASIVA DE COMBUSTIBLE")
    titulo.font = Font(bold=True, size=13, color=BRAND)
    ins.column_dimensions["A"].width = 110
    for r, linea in enumerate(INSTRUCCIONES, start=3):
        ins.cell(row=r, column=1, value=f"• {linea}")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --------------------------------------------------------------------------
# Parseo (estructural, sin validación de negocio)
# --------------------------------------------------------------------------


def _normaliza(texto: str) -> str:
    """minúsculas, sin acentos, sin '*', sin '(...)', espacios colapsados."""
    s = unicodedata.normalize("NFD", texto)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = s.lower().replace("*", " ")
    while "(" in s:
        a = s.index("(")
        b = s.find(")", a)
        s = s[:a] + (" " + s[b + 1 :] if b != -1 else " ")
    return " ".join(s.split())


# Encabezado normalizado -> campo. Tolera variantes razonables.
_ALIAS = {
    "matricula": "matricula",
    "fecha de carga": "fecha",
    "fecha": "fecha",
    "hora": "hora",
    "litros": "litros",
    "monto total": "monto",
    "monto": "monto",
    "importe": "monto",
    "moneda": "moneda",
    "tipo de cambio": "tipo_cambio",
    "tc": "tipo_cambio",
    "tipo combustible": "tipo_combustible",
    "tipo de combustible": "tipo_combustible",
    "combustible": "tipo_combustible",
    "lugar": "lugar",
    "aeropuerto": "lugar",
    "proveedor": "proveedor",
    "medio de pago": "medio_pago",
    "medio pago": "medio_pago",
    "metodo de pago": "medio_pago",
    "folio vuelo": "folio_vuelo",
    "folio de vuelo": "folio_vuelo",
    "folio": "folio_vuelo",
    "comprobante": "comprobante",
    "notas": "notas",
    "nota": "notas",
    "observaciones": "notas",
}


def _texto(v: object) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _fecha(v: object) -> str | None:
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return _texto(v)


def _hora(v: object) -> str | None:
    if isinstance(v, datetime):
        return v.strftime("%H:%M")
    if isinstance(v, time):
        return v.strftime("%H:%M")
    if isinstance(v, (int, float)) and not isinstance(v, bool) and 0 <= v < 1:
        minutos = round(float(v) * 24 * 60)  # fracción de día de Excel
        return f"{minutos // 60 % 24:02d}:{minutos % 60:02d}"
    return _texto(v)


def _numero(v: object) -> float | str | None:
    if v is None:
        return None
    if isinstance(v, bool):
        return _texto(v)
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    limpio = s.replace("$", "").replace(",", "").replace(" ", "")
    try:
        return float(limpio)
    except ValueError:
        return s  # crudo: el API lo reporta como ilegible


def _folio(v: object) -> str | None:
    if isinstance(v, bool):
        return _texto(v)
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    if isinstance(v, int):
        return str(v)
    return _texto(v)


def _leer_xlsx(raw: bytes) -> tuple[list, list[tuple[int, list]]]:
    try:
        wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    except Exception as e:  # zip corrupto, formato viejo, etc.
        raise ValueError(
            "No se pudo leer el archivo Excel. Sube el .xlsx de la plantilla "
            "de combustible (o un .csv)."
        ) from e
    try:
        objetivo = _normaliza(HOJA_CARGAS)
        nombre = next(
            (n for n in wb.sheetnames if _normaliza(n) == objetivo), wb.sheetnames[0]
        )
        ws = wb[nombre]
        encabezado: list | None = None
        filas: list[tuple[int, list]] = []
        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if encabezado is None:
                if any(c is not None and str(c).strip() for c in row):
                    encabezado = list(row)
                continue
            filas.append((idx, list(row)))
    finally:
        wb.close()
    if encabezado is None:
        raise ValueError("El archivo está vacío: no se encontró la fila de encabezados.")
    return encabezado, filas


def _leer_csv(raw: bytes) -> tuple[list, list[tuple[int, list]]]:
    try:
        texto = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        texto = raw.decode("latin-1")
    try:
        dialecto: type[csv.Dialect] | csv.Dialect = csv.Sniffer().sniff(
            texto[:4096], delimiters=",;\t"
        )
    except csv.Error:
        dialecto = csv.excel
    encabezado: list | None = None
    filas: list[tuple[int, list]] = []
    for idx, row in enumerate(csv.reader(io.StringIO(texto), dialecto), start=1):
        if encabezado is None:
            if any(str(c).strip() for c in row):
                encabezado = list(row)
            continue
        filas.append((idx, list(row)))
    if encabezado is None:
        raise ValueError("El CSV está vacío: no se encontró la fila de encabezados.")
    return encabezado, filas


def parse_combustible(req: ParseCombustibleRequest) -> ParseCombustibleResponse:
    try:
        raw = base64.b64decode(req.archivo_base64, validate=True)
    except (binascii.Error, ValueError) as e:
        raise ValueError("El archivo no llegó completo (base64 inválido).") from e
    if not raw:
        raise ValueError("El archivo está vacío.")

    if (req.filename or "").lower().endswith(".csv"):
        encabezado, filas = _leer_csv(raw)
    else:
        encabezado, filas = _leer_xlsx(raw)

    # Mapa campo -> índice de columna, por ENCABEZADO (no por posición).
    indices: dict[str, int] = {}
    for i, celda in enumerate(encabezado):
        if celda is None:
            continue
        campo = _ALIAS.get(_normaliza(str(celda)))
        if campo and campo not in indices:
            indices[campo] = i
    if "matricula" not in indices and "monto" not in indices:
        raise ValueError(
            "No se reconocieron los encabezados. ¿Es la plantilla de "
            "'Cargas de combustible'? No cambies los títulos de las columnas."
        )

    def _val(valores: list, campo: str) -> object:
        i = indices.get(campo)
        if i is None or i >= len(valores):
            return None
        return valores[i]

    resultado: list[FilaCombustible] = []
    for num, valores in filas:
        if not any(v is not None and str(v).strip() for v in valores):
            continue  # fila totalmente vacía
        matricula = _texto(_val(valores, "matricula"))
        monto = _numero(_val(valores, "monto"))
        if matricula is None and monto is None:
            continue  # sin matrícula ni monto no hay carga que evaluar
        resultado.append(
            FilaCombustible(
                fila=num,
                matricula=matricula,
                fecha=_fecha(_val(valores, "fecha")),
                hora=_hora(_val(valores, "hora")),
                litros=_numero(_val(valores, "litros")),
                monto=monto,
                moneda=_texto(_val(valores, "moneda")),
                tipo_cambio=_numero(_val(valores, "tipo_cambio")),
                tipo_combustible=_texto(_val(valores, "tipo_combustible")),
                lugar=_texto(_val(valores, "lugar")),
                proveedor=_texto(_val(valores, "proveedor")),
                medio_pago=_texto(_val(valores, "medio_pago")),
                folio_vuelo=_folio(_val(valores, "folio_vuelo")),
                comprobante=_texto(_val(valores, "comprobante")),
                notas=_texto(_val(valores, "notas")),
            )
        )
    return ParseCombustibleResponse(filas=resultado)
