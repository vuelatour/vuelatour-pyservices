"""Reporte consolidado de UN vuelo en Excel (openpyxl), por secciones.

El layout replica el formato de los controles del equipo ("Balance VGV.xlsx" /
"Dinero <mes>.xlsx"): doble moneda (USD y MXN con el tipo de cambio explícito),
IVA separado, venta por hora, tacómetro inicio/fin, gasavión/turbosina con
litros y $/litro, y el balance del vuelo (remanente → ganancia → ganancia x hr
→ % de ganancia). Los montos vienen YA calculados del API (aquí solo se
muestran; a lo sumo se multiplica por el TC para la columna en pesos).
"""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from app.schemas.reportes import ReporteVueloParticipacion, ReporteVueloRequest
from app.services.tabla_xlsx import sheet_title

_CANCUN = ZoneInfo("America/Cancun")


def _fecha(s: str | None) -> str:
    """ISO → dd/mm/aaaa HH:MM en hora Cancún (fechas puras se muestran tal cual)."""
    if not s:
        return "—"
    try:
        if len(s) == 10:  # date puro (fecha_gasto/fecha de cobro)
            d = datetime.fromisoformat(s)
            return d.strftime("%d/%m/%Y")
        dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=ZoneInfo("UTC"))
        return dt.astimezone(_CANCUN).strftime("%d/%m/%Y %H:%M")
    except ValueError:
        return s

def _pct(factor: float | None) -> str:
    """0.5 → '50 %' (hasta 2 decimales, sin ceros de más)."""
    return "—" if factor is None else f"{round(factor * 100, 2):g} %"


def _tramos_txt(tramos) -> str:
    """Tramos del avión en un vuelo multi-avión: el API puede mandar la
    cuenta (int), un texto ("CUN→MID") o una lista de órdenes."""
    if tramos is None or tramos == "" or tramos == []:
        return ""
    if isinstance(tramos, (list, tuple)):
        return "tramos " + ", ".join(str(t) for t in tramos)
    if isinstance(tramos, int):
        return f"{tramos} tramo{'s' if tramos != 1 else ''}"
    return str(tramos)


def _participacion_txt(p: ReporteVueloParticipacion) -> str:
    """'N990GG 50 % (CUN→MID) = $1,000.00 USD' — una matrícula del reparto."""
    tramos = _tramos_txt(p.tramos)
    venta = f" = ${p.venta_usd:,.2f} USD" if p.venta_usd is not None else ""
    return f"{p.matricula or '—'} {_pct(p.factor)}" + (f" ({tramos})" if tramos else "") + venta


BRAND = "DC2626"
NAVY = "102A43"
LIGHT = "F0F4F8"
GREEN = "15803D"
MONEY = '"$"#,##0.00'
PCT = "0.0%"
MUTED = "627D98"
N_COLS = 8


def _pago_vendedor(r: ReporteVueloRequest) -> float:
    """Pago al vendedor = comisión + su IVA cuando grava (`pagoVendedorUsd`,
    fuente única del API). API viejo sin el campo: la comisión pre-IVA."""
    return (
        r.pago_vendedor_usd
        if r.pago_vendedor_usd is not None
        else r.comision_vendedor_usd
    )


def _etiqueta_pago_vendedor(r: ReporteVueloRequest) -> str:
    quien = f" ({r.comision_vendedor_nombre})" if r.comision_vendedor_nombre else ""
    con_iva = " (c/IVA)" if _pago_vendedor(r) > r.comision_vendedor_usd + 0.005 else ""
    return f"(−) Pago comisión vendedor{quien}{con_iva}"


def render_reporte_vuelo_xlsx(r: ReporteVueloRequest) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title(f"Vuelo {r.folio}")
    ws.column_dimensions["A"].width = 26
    for col in ("B", "C", "D", "E", "F", "G", "H"):
        ws.column_dimensions[col].width = 15

    row = 1
    tc = r.tc_usd_mxn if (r.tc_usd_mxn or 0) > 0 else None

    def titulo(texto: str) -> None:
        nonlocal row
        c = ws.cell(row=row, column=1, value=texto)
        c.font = Font(bold=True, color="FFFFFF", size=12)
        c.fill = PatternFill("solid", fgColor=BRAND)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=N_COLS)
        row += 1

    def kv(label: str, value) -> None:
        nonlocal row
        ws.cell(row=row, column=1, value=label).font = Font(color=MUTED)
        ws.cell(row=row, column=2, value=value)
        row += 1

    def header(cols: list[str]) -> None:
        nonlocal row
        for i, h in enumerate(cols, start=1):
            c = ws.cell(row=row, column=i, value=h)
            c.font = Font(bold=True, color=NAVY, size=9)
            c.fill = PatternFill("solid", fgColor=LIGHT)
            c.alignment = Alignment(wrap_text=True, vertical="top")
        row += 1

    def money_cell(rr: int, cc: int, v, bold: bool = False, color: str | None = None) -> None:
        cell = ws.cell(row=rr, column=cc, value=v)
        cell.number_format = MONEY
        cell.alignment = Alignment(horizontal="right")
        if bold or color:
            cell.font = Font(bold=bold, color=color or "000000")

    # ===== Encabezado =====
    h = ws.cell(row=row, column=1, value=f"Reporte de vuelo #{r.folio}")
    h.font = Font(bold=True, size=14, color=BRAND)
    row += 1
    ws.cell(
        row=row, column=1, value=f"Generado {_fecha(r.generado)} · hora de Cancún (UTC-5)"
    ).font = Font(
        color=MUTED, italic=True
    )
    row += 2

    # ===== Resumen =====
    titulo("Datos del vuelo")
    kv("Cliente", r.cliente or "—")
    kv("Fecha de vuelo", _fecha(r.fecha_vuelo))
    kv("Ruta", r.ruta or "—")
    kv("Aeronave", r.aeronave or "—")
    # Vuelo MULTI-AVIÓN (regla B, 28-ago-2026): venta del avión repartida por
    # tramo entre las matrículas (gastos al avión de su tramo; TUAs/extras/
    # pernocta/comisión del vendedor son de VuelaTour, no se reparten).
    if r.participacion_aviones and len(r.participacion_aviones) > 1:
        kv(
            "Venta del avión por tramo (partes iguales por tramo vendido)",
            " · ".join(_participacion_txt(p) for p in r.participacion_aviones),
        )
    pil = r.piloto or "—"
    if r.copiloto:
        pil += f" / {r.copiloto} (copiloto)"
    kv("Piloto", pil)
    kv("Pasajeros", r.pasajeros)
    if r.pasajeros_nombres:
        kv("Nombres pasajeros", r.pasajeros_nombres)
    kv("Tipo / Estado", f"{r.tipo} · {r.estado}".strip(" ·"))
    if r.fecha_traslado_final:
        kv("Traslado final", _fecha(r.fecha_traslado_final))
    row += 1

    # ===== Venta (formato del control del equipo: hora, s/IVA, IVA, TC, pesos) =====
    titulo("Venta (cobrado al cliente)")
    header(
        [
            "TIEMPO\nCOBRADO HR",
            "VENTA HR\nS/IVA USD",
            "VENTA HR\nS/IVA MXN",
            "TOTAL\nS/IVA USD",
            "IVA\nUSD",
            "TOTAL\nUSD",
            "TIPO DE\nCAMBIO",
            "TOTAL\nMXN",
        ]
    )
    ws.cell(row=row, column=1, value=r.tiempo_cobrable_hr)
    money_cell(row, 2, r.tarifa_hora_usd or 0)
    if tc and r.tarifa_hora_usd:
        money_cell(row, 3, round(r.tarifa_hora_usd * tc, 2))
    money_cell(row, 4, r.venta_sin_iva_usd or max(r.total_usd - r.iva_usd, 0))
    money_cell(row, 5, r.iva_usd)
    money_cell(row, 6, r.total_usd, bold=True)
    if tc:
        ws.cell(row=row, column=7, value=tc).alignment = Alignment(horizontal="right")
        money_cell(row, 8, r.total_mxn if r.total_mxn else round(r.total_usd * tc, 2), bold=True)
    row += 2

    # Desglose canónico v1.3: las líneas SUMAN el total exacto — se imprimen
    # TODAS (aunque valgan 0), no omitir líneas del desglose.
    ws.cell(row=row, column=1, value="Desglose de la cotización").font = Font(bold=True, color=NAVY)
    row += 1
    for label, val in [
        ("Subtotal vuelo USD", r.subtotal_usd),
        ("TUAS USD", r.tuas_usd),
        ("Pernocta USD", r.viaticos_pernocta_usd),
        ("Extras USD", r.extras_total_usd),
        ("Ajuste USD", r.ajuste_final_usd),
        ("IVA USD", r.iva_usd),
    ]:
        ws.cell(row=row, column=1, value=label).font = Font(color=MUTED)
        money_cell(row, 2, val)
        row += 1
        # Detalle de TUAS por aeropuerto CON su moneda (requisito del cliente).
        # Sub-filas INFORMATIVAS (el monto viene en el propio texto): la fila
        # numérica "TUAS USD" de arriba sigue cuadrando la suma del desglose.
        if label == "TUAS USD" and r.tuas_detalle:
            for det in r.tuas_detalle:
                c = ws.cell(row=row, column=1, value=f"    {det}")
                c.font = Font(color=MUTED, size=9, italic=True)
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=N_COLS)
                row += 1
    ws.cell(row=row, column=1, value="Total USD").font = Font(bold=True)
    money_cell(row, 2, r.total_usd, bold=True)
    row += 1
    # Regla 28-ago-2026 (informativo, solo si el API lo manda): del total, la
    # VENTA DEL AVIÓN (tiempo + ajuste + IVA proporcional) y el ingreso de
    # VuelaTour (TUAs/extras/pernocta/comisión del vendedor + su IVA).
    for label, val in (
        ("    De esto, venta del avión USD", r.venta_avion_usd),
        (
            "    Ingreso VuelaTour (TUAs/extras/pernocta/comisión vendedor) USD",
            r.otros_ingresos_vuelatour_usd,
        ),
    ):
        if val is not None:
            ws.cell(row=row, column=1, value=label).font = Font(color=MUTED, size=9, italic=True)
            money_cell(row, 2, val, color=MUTED)
            row += 1
    # Comisión del vendedor (interna): el cliente paga el total completo; el
    # NETO VuelaTour (total − pago al vendedor) — distinto de la GANANCIA del
    # balance (que además resta gastos). Regla A (28-ago-2026): la comisión
    # es ingreso de VuelaTour y su pago sale de VuelaTour (no del avión);
    # este reporte mira la economía COMPLETA del vuelo y la resta UNA vez.
    if r.comision_vendedor_usd:
        quien = f" ({r.comision_vendedor_nombre})" if r.comision_vendedor_nombre else ""
        # Se resta el PAGO al vendedor (comisión + su IVA cuando grava), no la
        # comisión pre-IVA: así Total − esta fila = Neto y la aritmética
        # visible cuadra (antes restaba 832.00 a la vista y 965.12 en el neto).
        pago = _pago_vendedor(r)
        con_iva = " (comisión + IVA)" if pago > r.comision_vendedor_usd + 0.005 else ""
        ws.cell(
            row=row, column=1, value=f"Pago comisión vendedor{quien}{con_iva}"
        ).font = Font(color=MUTED)
        money_cell(row, 2, -pago)
        row += 1
        # Neto = total − PAGO al vendedor (comisión + su IVA), misma regla
        # que el balance de abajo.
        neto = (
            r.neto_vuelatour_usd
            if r.neto_vuelatour_usd is not None
            else r.total_usd - pago
        )
        ws.cell(row=row, column=1, value="Neto VuelaTour USD").font = Font(bold=True)
        money_cell(row, 2, neto, bold=True)
        row += 1
    if r.tarifa_tipo:
        kv("Tarifa", r.tarifa_tipo)
    if r.metodo_cobro:
        kv("Método de cobro", r.metodo_cobro)
    row += 1

    # ===== Tacómetro =====
    titulo("Tacómetro")
    # También con payload sin taco_inicio/fin (API viejo): la comparación de
    # horas cotizadas vs voladas no debe perderse.
    if (
        r.taco_inicio is not None
        or r.taco_fin is not None
        or r.horas_voladas_hr is not None
        or r.horas_cotizadas_hr is not None
    ):
        header(["TACO INICIO", "TACO FINAL", "HORAS VOLADAS", "HORAS COTIZADAS", "DIFERENCIA"])
        ws.cell(row=row, column=1, value=r.taco_inicio)
        ws.cell(row=row, column=2, value=r.taco_fin)
        ws.cell(row=row, column=3, value=r.horas_voladas_hr)
        ws.cell(row=row, column=4, value=r.horas_cotizadas_hr)
        ws.cell(row=row, column=5, value=r.horas_delta_hr)
        row += 2
    header(["#", "Tramo", "Salida", "Llegada", "Horas"])
    for t in r.tramos:
        ws.cell(row=row, column=1, value=t.orden)
        ws.cell(row=row, column=2, value=t.ruta)
        ws.cell(row=row, column=3, value=t.taco_salida)
        ws.cell(row=row, column=4, value=t.taco_llegada)
        ws.cell(row=row, column=5, value=t.horas)
        row += 1
    for nota in r.notas_horas:
        ws.cell(row=row, column=1, value=nota).font = Font(color=MUTED, size=9)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=N_COLS)
        row += 1
    row += 1

    # ===== Gasavión / Turbosina: combustible del AVIÓN (litros y $/litro,
    # como el control del equipo). No confundir con la categoría GASOLINA
    # (vehículos), que es un gasto común. =====
    titulo("Gasavión / Turbosina")
    if r.combustible:
        header(["Fecha", "Detalle", "Litros", "$ x litro", "Moneda", "Total"])
        for c in r.combustible:
            ws.cell(row=row, column=1, value=_fecha(c.fecha))
            ws.cell(row=row, column=2, value=c.detalle or c.concepto or "")
            if c.litros:
                ws.cell(row=row, column=3, value=c.litros)
                if c.monto:
                    money_cell(row, 4, round(c.monto / c.litros, 2))
            ws.cell(row=row, column=5, value=c.moneda or "")
            money_cell(row, 6, c.monto)
            row += 1
        if r.combustible_total_usd:
            tot = ws.cell(row=row, column=1, value="Total gasavión / turbosina USD")
            tot.font = Font(bold=True)
            money_cell(row, 2, r.combustible_total_usd, bold=True)
            row += 1
    else:
        ws.cell(row=row, column=1, value="Sin cargas de combustible registradas.").font = Font(
            color=MUTED
        )
        row += 1
    row += 1

    # ===== Gastos =====
    titulo("Gastos del vuelo")
    if r.gastos:
        header(["Fecha", "Categoría", "Detalle", "Moneda", "Monto"])
        for g in r.gastos:
            ws.cell(row=row, column=1, value=_fecha(g.fecha))
            # Categoría: etiqueta amable si el API la manda; si no, el código.
            ws.cell(row=row, column=2, value=g.etiqueta or g.concepto or "")
            ws.cell(row=row, column=3, value=g.detalle or "")
            ws.cell(row=row, column=4, value=g.moneda or "")
            money_cell(row, 5, g.monto)
            row += 1
        if r.gastos_total_usd:
            ws.cell(row=row, column=1, value="Total gastos USD").font = Font(bold=True)
            money_cell(row, 2, r.gastos_total_usd, bold=True)
            row += 1
    else:
        ws.cell(row=row, column=1, value="Sin gastos registrados.").font = Font(color=MUTED)
        row += 1
    if r.gastos_sin_tc_count:
        ws.cell(
            row=row,
            column=1,
            value=(
                f"OJO: {r.gastos_sin_tc_count} gasto(s) en MXN por ${r.gastos_sin_tc_mxn:,.2f} "
                "SIN tipo de cambio: no entran al balance USD."
            ),
        ).font = Font(color=BRAND, size=9)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=N_COLS)
        row += 1
    row += 1

    # ===== Status de cobros =====
    titulo("Status de cobros")
    if r.cobros:
        header(["Fecha", "Método", "Moneda", "Cantidad"])
        for c in r.cobros:
            ws.cell(row=row, column=1, value=_fecha(c.fecha))
            ws.cell(row=row, column=2, value=c.concepto or "")
            ws.cell(row=row, column=3, value=c.moneda or "")
            money_cell(row, 4, c.monto)
            row += 1
            # Comisión bancaria del cobro (el banco depositó monto − comisión).
            if c.detalle:
                ws.cell(row=row, column=2, value=c.detalle).font = Font(color=MUTED, size=9)
                row += 1
    else:
        ws.cell(row=row, column=1, value="Sin cobros registrados.").font = Font(color=MUTED)
        row += 1
    ws.cell(row=row, column=1, value="Total cobrado USD").font = Font(bold=True)
    money_cell(row, 2, r.total_cobrado_usd, bold=True)
    row += 1
    if r.comision_banco_usd:
        ws.cell(row=row, column=1, value="Comisiones bancarias").font = Font(color=MUTED)
        money_cell(row, 2, -r.comision_banco_usd)
        row += 1
        neto_banco = (
            r.total_cobrado_neto_usd
            if r.total_cobrado_neto_usd is not None
            else r.total_cobrado_usd - r.comision_banco_usd
        )
        ws.cell(row=row, column=1, value="Neto recibido (después de comisión)").font = Font(
            bold=True
        )
        money_cell(row, 2, neto_banco, bold=True)
        row += 1
    # "ME DEBEN", como la columna del control del equipo.
    ws.cell(row=row, column=1, value="ME DEBEN (saldo del cliente)").font = Font(
        bold=True, color=BRAND if (r.saldo_usd or 0) > 0 else GREEN
    )
    money_cell(row, 2, r.saldo_usd, bold=True, color=BRAND if (r.saldo_usd or 0) > 0 else GREEN)
    row += 2

    # ===== Balance del vuelo (remanente → ganancia → % — corazón del Excel
    # del equipo). La columna MXN es el equivalente al TC del vuelo; se deriva
    # EN CADENA de las filas redondeadas para que la columna sume exacta.
    # Solo con datos de economía (payload viejo / RESERVA sin cotizar: nada). =====
    if r.remanente_usd is not None or r.ganancia_final_usd is not None:
        titulo("Balance del vuelo")
        header(["Concepto", "USD", f"MXN (T.C. {tc})" if tc else "MXN"])

        def bal(
            label: str,
            usd,
            bold: bool = False,
            color: str | None = None,
            signo: int = 1,
            mxn: float | None = None,
        ) -> float:
            nonlocal row
            ws.cell(row=row, column=1, value=label).font = Font(
                bold=bold, color=color or ("000000" if bold else MUTED)
            )
            mxn_val = 0.0
            if usd is not None:
                money_cell(row, 2, signo * usd, bold=bold, color=color)
                if tc is not None:
                    mxn_val = mxn if mxn is not None else round(usd * tc, 2)
                    money_cell(row, 3, round(signo * mxn_val, 2), bold=bold, color=color)
            row += 1
            return mxn_val

        venta_mxn = bal("Venta total (c/IVA)", r.total_usd, bold=True)
        bal("Venta sin IVA", r.venta_sin_iva_usd or max(r.total_usd - r.iva_usd, 0))
        gas_mxn = bal("(−) Gasavión / Turbosina", r.combustible_total_usd or 0, signo=-1)
        gtos_mxn = bal("(−) Gastos del vuelo", r.gastos_total_usd or 0, signo=-1)
        rem_mxn = 0.0
        if r.remanente_usd is not None:
            rem_mxn = bal(
                "REMANENTE (venta − costo)",
                r.remanente_usd,
                bold=True,
                mxn=round(venta_mxn - gas_mxn - gtos_mxn, 2),
            )
        comv_mxn = comb_mxn = 0.0
        if _pago_vendedor(r):
            # Pago de VuelaTour al vendedor (regla A): se resta UNA vez, con
            # su IVA cuando grava (pago_vendedor_usd del API).
            comv_mxn = bal(_etiqueta_pago_vendedor(r), _pago_vendedor(r), signo=-1)
        if r.comision_banco_usd:
            comb_mxn = bal("(−) Comisiones bancarias", r.comision_banco_usd, signo=-1)
        if r.ganancia_final_usd is not None:
            color = GREEN if r.ganancia_final_usd >= 0 else BRAND
            bal(
                "GANANCIA DEL VUELO",
                r.ganancia_final_usd,
                bold=True,
                color=color,
                mxn=(
                    round(rem_mxn - comv_mxn - comb_mxn, 2)
                    if r.remanente_usd is not None
                    else None
                ),
            )
        if r.ganancia_x_hr_usd is not None:
            bal("Ganancia x hora", r.ganancia_x_hr_usd)
        if r.ganancia_pct is not None:
            ws.cell(row=row, column=1, value="% de ganancia (sobre venta s/IVA)").font = Font(
                bold=True
            )
            c = ws.cell(row=row, column=2, value=r.ganancia_pct)
            c.number_format = PCT
            c.alignment = Alignment(horizontal="right")
            c.font = Font(bold=True, color=GREEN if r.ganancia_pct >= 0 else BRAND)
            row += 1

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
