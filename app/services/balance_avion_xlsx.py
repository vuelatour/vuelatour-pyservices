"""Balance mensual por avión en Excel (openpyxl) — réplica sistematizada del
control del equipo ("Balance N990GG.xlsx").

Ocho hojas en el orden del libro:
  1. reporte horas <MATRÍCULA> — maestro: 1 fila = 1 vuelo, TOTALES al final.
     VENTA AVIÓN = tiempo de vuelo + ajuste + IVA proporcional (regla
     28-ago-2026: TUAs/extras/pernocta cobrados son ingreso de VuelaTour y
     viven en 'otros movimientos' del Balance general). Costos SIN
     combustible; el TUA pagado es SOLO nota en OPERACIONES (no resta).
     STATUS DE COBROS trae los depósitos REALES (COBRO 1..4 y su Σ) y,
     aparte, COBRADO AVIÓN = la parte prorrateada al avión.
  2. cobranza — estatus de cobro por vuelo: venta avión prorrateada, total
     cotización (c/extras), depósitos reales, comisión y cuenta del banco.
  3. combustible — el gas del avión POR MES (litros y $/L), 26-ago-2026.
  4. gastos indirectos (gastos del avión sin vuelo)  5. otros gastos (parte
     de este avión de los gastos administrativos repartidos a mano)
  6. permisos (todo AFAC).
  7. balance — cascada (− combustible − indirectos − otros − permisos) +
     reparto real de socios.
  8. pendientes de captura — lo que falta para que el libro quede completo.

Los montos vienen YA calculados del API (aquí solo se pintan; jamás se
recalcula dinero). None = celda vacía — nunca un 0 falso.
"""

from __future__ import annotations

import math
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.schemas.reportes import (
    BalanceAvionCobro,
    BalanceAvionHojaCombustible,
    BalanceAvionHojaGastos,
    BalanceAvionRequest,
    BalanceAvionVuelo,
    BalanceGeneralRequest,
    BalanceGeneralResumenFila,
    BalanceHojaOtrosMovimientos,
    BalanceOtroMovimientoFila,
)
from app.services.tabla_xlsx import sheet_title

BRAND = "0F4C81"
NAVY = "102A43"
MUTED = "627D98"
LIGHT = "EEF2F7"
GREEN = "15803D"
RED = "DC2626"
AMBER = "FCD34D"  # horas cobradas < voladas (regla: nunca cobrar de menos)
# TC oficial autocompletado (Banxico FIX): azul claro, sutil (pedido 27-ago).
TC_OFICIAL_FILL = "DCE9F8"
# Colores suaves por bloque (ayuda visual pedida en el contrato).
FILL_VENTA = "E8F0FB"  # azul suave
FILL_COSTOS = "FDF3E7"  # ámbar suave
FILL_COBROS = "E9F6EC"  # verde suave

MONEY = "#,##0.00"
HORAS = "0.00"
TACO = "0.0"
TC = "0.0000"

_thin = Side(style="thin", color="D5DBE3")
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _hex(color: str | None) -> str | None:
    """#RRGGBB → RRGGBB tal cual (mismo criterio que el Libro Dinero: el
    equipo usa sus colores saturados por avión, no se aclaran)."""
    if not color:
        return None
    c = color.lstrip("#").strip()
    return c.upper() if len(c) == 6 else None


def _fecha(s: str | None) -> str | None:
    """ISO date → dd/mm/aaaa; texto libre (multi-día '9-10 sep') tal cual."""
    if not s:
        return None
    try:
        if len(s) == 10:
            return datetime.fromisoformat(s).strftime("%d/%m/%Y")
        return datetime.fromisoformat(s.replace("Z", "+00:00")).strftime("%d/%m/%Y")
    except ValueError:
        return s


def _title(ws: Worksheet, text: str, row: int, span: int, size: int = 14) -> None:
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(bold=True, size=size, color=BRAND)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)


def _header_row(ws: Worksheet, row: int, headers: list[str], start: int = 1) -> None:
    for col, h in enumerate(headers, start=start):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=BRAND)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border


def _num(ws: Worksheet, row: int, col: int, value: float | None, fmt: str = MONEY, **font):
    """Pinta un numérico; None → celda vacía (nunca 0 falso)."""
    cell = ws.cell(row=row, column=col)
    cell.number_format = fmt
    cell.alignment = Alignment(horizontal="right")
    if value is not None:
        cell.value = value
    if font:
        cell.font = Font(**font)
    return cell


# ===== Hoja maestra: definición de columnas (orden del Excel original) =====
# (grupo, encabezado, atributo del vuelo, formato numérico o None=texto)
_COLS: list[tuple[str, str, str | None, str | None]] = [
    ("", "CLAVE", "clave", None),
    ("", "FECHA", "fecha", None),
    ("", "RUTA", "ruta", None),
    ("", "ESTADO", "estado", None),
    ("VENTA", "HORAS\nCOBRADAS", "horas_cobradas", HORAS),
    ("VENTA", "TARIFA\nUSD/HR S/IVA", "tarifa_usd", MONEY),
    ("VENTA", "IVA\nUSD/HR", "iva_hr_usd", MONEY),
    ("VENTA", "VENTA AVIÓN\nUSD *", "total_usd", MONEY),
    ("VENTA", "IVA VENTA\nAVIÓN USD", "iva_usd", MONEY),
    ("VENTA", "TIPO CAMBIO\nVENTA", "tc_venta", TC),
    ("VENTA", "VENTA AVIÓN\nMXN", "total_mxn", MONEY),
    ("VENTA", "IVA VENTA\nAVIÓN MXN", "iva_mxn", MONEY),
    ("VENTA", "TOTAL S/IVA\nMXN", "subtotal_mxn", MONEY),
    ("TIEMPO / TACÓMETRO", "TIEMPO\nVUELO HR", "tiempo_vuelo", HORAS),
    ("TIEMPO / TACÓMETRO", "TACO\nINICIO", "taco_inicio", TACO),
    ("TIEMPO / TACÓMETRO", "TACO\nFINAL", "taco_fin", TACO),
    # GAS fuera de la hoja maestra (26-ago-2026): el combustible se controla
    # por avión/mes en su propia hoja "combustible" (los campos gas_* siguen
    # en el esquema por compat de contrato; un API viejo aún los manda).
    ("COSTOS DIRECTOS (MXN)", "OPERACIONES", "op_mxn", MONEY),
    ("COSTOS DIRECTOS (MXN)", "PILOTO", "piloto_mxn", MONEY),
    ("COSTOS DIRECTOS (MXN)", "OTROS", "otros_mxn", MONEY),
    ("COSTOS DIRECTOS (MXN)", "PERMISO AFAC\n(PROVISIÓN)", "permiso_afac_mxn", MONEY),
    ("COSTOS DIRECTOS (MXN)", "COSTO TOTAL\nMXN", "costo_total_mxn", MONEY),
    ("COSTOS DIRECTOS (MXN)", "TIPO CAMBIO\nCOSTOS", "tc_costos", TC),
    ("INDICADORES USD / IVA", "COSTO TOTAL\nUSD", "costo_usd", MONEY),
    ("INDICADORES USD / IVA", "COSTO TOTAL\nUSD S/IVA", "costo_usd_siva", MONEY),
    ("INDICADORES USD / IVA", "IVA PAGADO\nUSD", "iva_pagado_usd", MONEY),
    ("INDICADORES USD / IVA", "IVA PAGADO\nMXN", "iva_pagado_mxn", MONEY),
    ("INDICADORES USD / IVA", "REMANENTE\nVENTA−COSTO MXN", "remanente_mxn", MONEY),
    ("INDICADORES USD / IVA", "DIF. IVA\nHACIENDA MXN", "dif_iva_mxn", MONEY),
    ("INDICADORES USD / IVA", "COMISIÓN\nVENDEDOR MXN", "comision_vendedor_mxn", MONEY),
    ("INDICADORES USD / IVA", "GANANCIA\nMXN", "ganancia_mxn", MONEY),
    ("INDICADORES USD / IVA", "GANANCIA\nUSD", "ganancia_usd", MONEY),
    ("INDICADORES USD / IVA", "COSTO X HORA\nUSD", "costo_hr_usd", MONEY),
    ("INDICADORES USD / IVA", "COSTO X HORA\nUSD S/IVA", "costo_hr_usd_siva", MONEY),
    ("STATUS DE COBROS", "STATUS", "status_cobro", None),
    ("STATUS DE COBROS", "COBRO 1\nFECHA", None, None),
    ("STATUS DE COBROS", "COBRO 1\nMXN", None, MONEY),
    ("STATUS DE COBROS", "COBRO 2\nFECHA", None, None),
    ("STATUS DE COBROS", "COBRO 2\nMXN", None, MONEY),
    ("STATUS DE COBROS", "COBRO 3\nFECHA", None, None),
    ("STATUS DE COBROS", "COBRO 3\nMXN", None, MONEY),
    ("STATUS DE COBROS", "COBRO 4\nFECHA", None, None),
    ("STATUS DE COBROS", "COBRO 4\nMXN", None, MONEY),
    # Regla 28-ago-2026: primero los depósitos REALES (Σ COBRO 1..4, tal
    # cual entraron) y aparte la parte de esos depósitos que corresponde
    # al AVIÓN (prorrateo hecho por el API; el resto es de VuelaTour).
    ("STATUS DE COBROS", "COBRADO REAL\nMXN (Σ depósitos)", "cobrado_real_mxn", MONEY),
    ("STATUS DE COBROS", "COBRADO AVIÓN\nMXN (prorrateado) ****", "cobrado_mxn", MONEY),
    ("STATUS DE COBROS", "POR COBRAR\nMXN", "por_cobrar_mxn", MONEY),
    ("STATUS DE COBROS", "POR COBRAR\nUSD", "por_cobrar_usd", MONEY),
]
_COBRO1_COL = next(i for i, c in enumerate(_COLS, start=1) if c[1] == "COBRO 1\nFECHA")
# Celdas de costos con NOTA de desglose (comentario de Excel): al pasar el
# cursor se ve qué gastos componen el total ("Comida · Starbucks — $206.00").
_DETALLE_ATTR = {
    "op_mxn": "op_detalle",
    "piloto_mxn": "piloto_detalle",
    "otros_mxn": "otros_detalle",
}
_GROUP_FILLS = {"VENTA": FILL_VENTA, "COSTOS DIRECTOS (MXN)": FILL_COSTOS,
                "STATUS DE COBROS": FILL_COBROS}
# Columnas de la fila TOTALES: atributo del vuelo → atributo de totales.
_TOTAL_MAP = {
    "horas_cobradas": "horas_cobradas", "tiempo_vuelo": "tiempo_vuelo",
    "total_mxn": "total_mxn", "iva_mxn": "iva_mxn", "subtotal_mxn": "subtotal_mxn",
    "op_mxn": "op_mxn",
    "piloto_mxn": "piloto_mxn", "otros_mxn": "otros_mxn",
    "permiso_afac_mxn": "permiso_afac_mxn", "costo_total_mxn": "costo_total_mxn",
    "remanente_mxn": "remanente_mxn", "dif_iva_mxn": "dif_iva_mxn",
    "comision_vendedor_mxn": "comision_vendedor_mxn", "ganancia_mxn": "ganancia_mxn",
    "ganancia_usd": "ganancia_usd", "cobrado_mxn": "cobrado_mxn",
    "cobrado_real_mxn": "cobrado_real_mxn",
    "por_cobrar_mxn": "por_cobrar_mxn", "por_cobrar_usd": "por_cobrar_usd",
    # Promedios (se pintan en su columna con etiqueta "prom." implícita):
    "tc_costos": "tc_promedio", "costo_hr_usd": "costo_hr_prom_usd",
}


def _cobros_a_4(cobros: list[BalanceAvionCobro]) -> list[BalanceAvionCobro]:
    """Hasta 4 parcialidades; si hay más, la 4ª agrega el resto (solo suma de
    columna para mostrar, no cálculo de negocio)."""
    if len(cobros) <= 4:
        return cobros
    resto = cobros[3:]
    montos = [c.monto_mxn for c in resto if c.monto_mxn is not None]
    agg = round(sum(montos), 2) if montos else None
    ultima = next((c.fecha for c in reversed(resto) if c.fecha), None)
    etiqueta = f"{_fecha(ultima)} (+{len(resto)})" if ultima else f"(+{len(resto)} cobros)"
    return [*cobros[:3], BalanceAvionCobro(fecha=etiqueta, monto_mxn=agg)]


def _hoja_maestra(ws: Worksheet, req: BalanceAvionRequest) -> None:
    ws.title = sheet_title(f"reporte horas {req.matricula}")
    n = len(_COLS)

    # Encabezado compacto de 2 filas: grupo (merged) / columna.
    col = 1
    while col <= n:
        grupo = _COLS[col - 1][0]
        fin = col
        while fin < n and _COLS[fin][0] == grupo:
            fin += 1
        if grupo:
            c = ws.cell(row=1, column=col, value=grupo)
            c.font = Font(bold=True, color="FFFFFF", size=10)
            c.fill = PatternFill("solid", fgColor=NAVY)
            c.alignment = Alignment(horizontal="center", vertical="center")
            if fin > col:
                ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=fin)
        col = fin + 1
    for i, (grupo, header, _attr, _fmt) in enumerate(_COLS, start=1):
        c = ws.cell(row=2, column=i, value=header)
        c.font = Font(bold=True, color=NAVY, size=8)
        c.fill = PatternFill("solid", fgColor=_GROUP_FILLS.get(grupo, LIGHT))
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
    ws.row_dimensions[2].height = 38

    # Datos: 1 fila por vuelo.
    row = 3
    for v in req.vuelos:
        cobros = _cobros_a_4(v.cobros)
        # Regla del cliente: lo cobrado nunca puede ser menor a lo volado.
        # Resalta HORAS COBRADAS en ámbar cuando el taco registró más horas
        # (solo señal visual; el pendiente del API explica el caso).
        horas_menores = (
            v.horas_cobradas is not None
            and v.tiempo_vuelo is not None
            and v.horas_cobradas > 0
            and v.tiempo_vuelo - v.horas_cobradas > 0.01
        )
        # Balance GENERAL: SOLO la celda de la CLAVE se tiñe con el color del
        # avión (así lo maneja el equipo en su libro — el resto de la fila
        # conserva los colores por bloque); en el individual viene vacío.
        avion_fill = _hex(v.avion_color)
        for i, (grupo, _header, attr, fmt) in enumerate(_COLS, start=1):
            fill = _GROUP_FILLS.get(grupo)
            if i == 1 and avion_fill:
                fill = avion_fill
            if attr is not None:
                val = getattr(v, attr)
                if fmt is None:
                    cell = ws.cell(row=row, column=i, value=_fecha(val) if attr == "fecha" else val)
                    if attr == "estado" and val == "CANCELADO":
                        cell.font = Font(color=RED, size=9)
                    elif attr == "estado":
                        cell.font = Font(color=MUTED, size=9)
                else:
                    cell = _num(ws, row, i, val, fmt)
                if attr == "horas_cobradas" and horas_menores:
                    fill = AMBER
                # TC no capturado en la cotización → se usó el oficial: la
                # celda del TC y las que derivan de él (MXN) se marcan.
                if v.tc_venta_oficial and attr in (
                    "tc_venta", "total_mxn", "iva_mxn", "subtotal_mxn",
                ):
                    fill = TC_OFICIAL_FILL
                    if attr == "tc_venta":
                        cell.comment = Comment(
                            "TC oficial (Banxico FIX / DOF) del día de la cotización: "
                            "la cotización no traía tipo de cambio.",
                            "VuelaTour",
                        )
                # Salto INTERNO entre tramos del MISMO vuelo: infla las horas
                # sin romper la cadena entre vuelos — se pinta en la celda de
                # horas voladas con el tramo culpable en la nota.
                if attr == "tiempo_vuelo" and v.salto_taco_interno:
                    fill = AMBER
                    nota_int = Comment(
                        "Salto entre tramos del vuelo (las horas pueden "
                        "salir infladas): "
                        + (v.salto_taco_interno_detalle or "revisar tacos"),
                        "VuelaTour",
                    )
                    nota_int.width = 280
                    nota_int.height = 70
                    cell.comment = nota_int
                # Salto en la cadena de tacómetros y/u OBSERVACIONES del
                # equipo (Tacómetros en vivo): mismo amarillo que el panel;
                # la nota de la celda junta el salto y los comentarios.
                if attr == "taco_inicio" and (
                    v.salto_taco_inicio or v.taco_inicio_obs
                ):
                    fill = AMBER
                    lineas_ti: list[str] = []
                    if v.salto_taco_inicio and v.salto_taco_esperado is not None:
                        lineas_ti.append(
                            "Salto en la cadena de tacómetros: no empalma "
                            f"con la llegada anterior ({v.salto_taco_esperado:g})."
                        )
                    elif v.salto_taco_inicio:
                        lineas_ti.append("Salto en la cadena de tacómetros.")
                    lineas_ti.extend(v.taco_inicio_obs)
                    if lineas_ti:
                        nota_salto = Comment("\n".join(lineas_ti), "VuelaTour")
                        nota_salto.width = 320
                        nota_salto.height = min(220, 45 + 30 * len(lineas_ti))
                        cell.comment = nota_salto
                # Observaciones sobre la LLEGADA → celda TACO FINAL.
                if attr == "taco_fin" and v.taco_fin_obs:
                    fill = AMBER
                    nota_tf = Comment("\n".join(v.taco_fin_obs), "VuelaTour")
                    nota_tf.width = 320
                    nota_tf.height = min(220, 45 + 30 * len(v.taco_fin_obs))
                    cell.comment = nota_tf
                # Nota con el desglose del total de la celda (gastos que la
                # componen), visible al pasar el cursor en Excel.
                det_attr = _DETALLE_ATTR.get(attr)
                if det_attr:
                    lineas = getattr(v, det_attr, None) or []
                    if lineas:
                        nota = Comment("\n".join(lineas), "VuelaTour")
                        nota.width = 340
                        nota.height = min(260, 40 + 16 * len(lineas))
                        cell.comment = nota
            else:  # parcialidades de cobro (pares fecha/monto desde _COBRO1_COL)
                idx = (i - _COBRO1_COL) // 2
                cobro = cobros[idx] if idx < len(cobros) else None
                if fmt is None:
                    cell = ws.cell(row=row, column=i, value=_fecha(cobro.fecha) if cobro else None)
                else:
                    cell = _num(ws, row, i, cobro.monto_mxn if cobro else None, fmt)
            cell.border = _border
            if fill:
                cell.fill = PatternFill("solid", fgColor=fill)
        row += 1

    # Fila TOTALES al final (sumas y promedios YA calculados por el API).
    t = req.totales
    ws.cell(row=row, column=1, value="TOTALES").font = Font(bold=True)
    for i, (_grupo, _header, attr, fmt) in enumerate(_COLS, start=1):
        cell = ws.cell(row=row, column=i)
        total_attr = _TOTAL_MAP.get(attr) if attr else None
        if total_attr is not None:
            cell = _num(ws, row, i, getattr(t, total_attr), fmt or MONEY, bold=True)
        cell.fill = PatternFill("solid", fgColor=LIGHT)
        cell.border = _border
        if cell.value is not None or i == 1:
            cell.font = Font(bold=True)
    row += 1

    # Renglones informativos (regla 28-ago-2026): lo que NO está en las filas.
    # 1) TUAs/extras/pernocta COTIZADOS (+ su IVA) = ingreso de VuelaTour,
    #    EXCLUIDOS de VENTA AVIÓN; el detalle vive en 'otros movimientos'
    #    del Balance general. Es lo cotizado de TODOS los estados (misma
    #    base que VENTA AVIÓN), no lo cobrado.
    otros = t.otros_ingresos_usd
    if otros is not None and otros != 0:
        lc = ws.cell(
            row=row,
            column=1,
            value="TUAs/extras/pernocta COTIZADOS en el periodo (con su IVA; "
            "todos los estados, igual que VENTA AVIÓN) — EXCLUIDOS de las "
            "filas: ingreso de VuelaTour, detalle en 'otros movimientos' "
            "del Balance general:",
        )
        lc.font = Font(bold=True, size=9)
        lc.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        ws.row_dimensions[row].height = 30
        _num(ws, row, 9, otros, MONEY, bold=True)
        ws.cell(row=row, column=10, value="USD").font = Font(bold=True, size=9)
        row += 1
    # 2) TUA pagado al aeropuerto: solo nota en OPERACIONES, no resta aquí.
    #    El API lo manda como suma → 0 = no hubo; se omite igual que None.
    tua = t.tua_pagado_mxn
    if tua is not None and tua != 0:
        ws.cell(
            row=row,
            column=1,
            value="TUA pagado del periodo (solo nota en OPERACIONES, no resta "
            "en este libro):",
        ).font = Font(bold=True, size=9)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        _num(ws, row, 9, tua, MONEY, bold=True)
        ws.cell(row=row, column=10, value="MXN").font = Font(bold=True, size=9)
        row += 1
    row += 1

    # Notas al pie.
    for nota in (
        "* VENTA AVIÓN de cada fila = tiempo de vuelo (tarifa × horas "
        "cobradas) + ajuste/descuento + su IVA proporcional. IVA VENTA "
        "AVIÓN (USD y MXN) es SOLO el IVA proporcional de la venta del "
        "avión — el IVA de TUAs/extras/pernocta viaja con ellos. TUAs, "
        "extras y viáticos de pernocta NO son venta del avión: son ingreso "
        "de VuelaTour (pestaña 'otros movimientos' del Balance general). "
        "Sin cotización: horas × tarifa.",
        "**** COBRADO REAL = Σ de los depósitos tal cual entraron (COBRO "
        "1..4). COBRADO AVIÓN = depósitos reales × (venta avión ÷ total "
        "cotización): la parte de los cobros que corresponde a TUAs/extras/"
        "pernocta es de VuelaTour (ver 'otros movimientos'). POR COBRAR "
        "es la parte del avión. COBRADO AVIÓN y POR COBRAR van al TC de "
        "venta (un depósito en pesos con su propio TC se pasa a USD con ese "
        "TC y se re-expresa al TC de venta), así que COBRADO AVIÓN puede "
        "diferir de COBRADO REAL sin que falte dinero.",
        "TIPO CAMBIO COSTOS y COSTO X HORA USD de la fila TOTALES son PROMEDIOS "
        "(los demás son sumas).",
        "El ESTATUS DE COBRO por vuelo (cuánto se cobró, con qué fechas y "
        "métodos, y cuánto falta) está al frente en la hoja 'cobranza' — el "
        "bloque STATUS DE COBROS del final de esta hoja trae lo mismo en "
        "columnas.",
        "El COMBUSTIBLE ya no va por vuelo (26-ago-2026): se controla por "
        "avión y por MES en la hoja 'combustible' (litros y $/L incluidos) y "
        "resta una sola vez en la hoja 'balance'. Por eso COSTO TOTAL, COSTO "
        "X HORA, REMANENTE y GANANCIA de las filas van SIN combustible (y el "
        "TUA pagado tampoco resta, ver **): la utilidad real del periodo se "
        "lee en la hoja 'balance', no en la fila. IVA PAGADO y DIF. IVA "
        "también derivan solo de los costos de la fila (sin gas ni TUA).",
        "** El TUA pagado al aeropuerto NO es costo del avión ni resta en "
        "ningún lado de este libro: queda solo como nota en la celda "
        "OPERACIONES ('TUA $x**'); cobro y pago del TUA viven en 'otros "
        "movimientos' del Balance general. Los servicios FBO sí son costo "
        "(columna OTROS).",
        "*** Filas 'COMPARTIDO': el vuelo mezcló aviones — aquí van SOLO los "
        "tramos, horas y costos de esta matrícula; la VENTA completa está en "
        "el balance del avión principal (el prorrateo del precio entre "
        "aviones está pendiente de decisión).",
        "TACO INICIO / TACO FINAL en ámbar = salto en la cadena de "
        "tacómetros (el valor esperado está en la nota de la celda) y/u "
        "OBSERVACIÓN del equipo capturada en Tacómetros en vivo — pasa el "
        "cursor por la celda para leer el comentario (quién y cuándo). "
        "TIEMPO VUELO en ámbar = salto entre tramos DEL MISMO vuelo (el "
        "tramo culpable está en la nota) — mismo amarillo que el panel.",
    ):
        ws.cell(row=row, column=1, value=nota).font = Font(color=MUTED, size=9, italic=True)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
        row += 1

    # Anchos + paneles congelados (bajo el encabezado, a la derecha de RUTA).
    anchos = {1: 16, 2: 12, 3: 34, 4: 12}
    for i, (_g, _h, attr, _f) in enumerate(_COLS, start=1):
        if attr in ("cobrado_real_mxn", "cobrado_mxn"):
            anchos[i] = 17
    for i in range(1, n + 1):
        ws.column_dimensions[get_column_letter(i)].width = anchos.get(i, 13)
    ws.freeze_panes = "D3"


def _hoja_gastos(ws: Worksheet, titulo: str, hoja: BalanceAvionHojaGastos,
                 req: BalanceAvionRequest, nota: str | None = None) -> None:
    _title(ws, f"{titulo} — {req.matricula}".strip(" —"), 1, 5)
    periodo = f"Periodo: {req.periodo_desde or '—'} a {req.periodo_hasta or '—'}"
    ws.cell(row=2, column=1, value=periodo).font = Font(italic=True, size=10, color=MUTED)

    _header_row(ws, 4, ["TOTAL MXN", "TC PROMEDIO", "TOTAL USD", "HRS VOLADAS", "USD X HR"])
    _num(ws, 5, 1, hoja.total_mxn, MONEY, bold=True)
    _num(ws, 5, 2, req.totales.tc_promedio, TC)
    _num(ws, 5, 3, hoja.usd, MONEY, bold=True)
    _num(ws, 5, 4, req.totales.tiempo_vuelo, HORAS)
    _num(ws, 5, 5, hoja.usd_hr, MONEY)
    for c in range(1, 6):
        ws.cell(row=5, column=c).border = _border

    _header_row(ws, 7, ["FECHA", "DETALLE", "MONTO MXN", "MONEDA ORIGINAL", "MONTO ORIGINAL"])
    row = 8
    if hoja.filas:
        for f in hoja.filas:
            ws.cell(row=row, column=1, value=_fecha(f.fecha)).border = _border
            ws.cell(row=row, column=2, value=f.detalle).border = _border
            _num(ws, row, 3, f.monto_mxn).border = _border
            # Moneda/monto original solo cuando el gasto NO se capturó en MXN.
            ws.cell(row=row, column=4, value=f.moneda_original).border = _border
            _num(ws, row, 5, f.monto_original).border = _border
            # Balance GENERAL: SOLO la celda del DETALLE (lleva la matrícula
            # al frente) se tiñe con el color del avión — como su libro.
            avion_fill = _hex(f.avion_color)
            if avion_fill:
                ws.cell(row=row, column=2).fill = PatternFill(
                    "solid", fgColor=avion_fill
                )
            row += 1
    else:
        ws.cell(row=row, column=1, value="Sin gastos registrados en el periodo.").font = Font(
            color=MUTED, italic=True
        )
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        row += 1

    if nota:
        row += 1
        ws.cell(row=row, column=1, value=nota).font = Font(
            color=MUTED, size=9, italic=True
        )
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)

    for i, w in enumerate([14, 52, 15, 16, 15], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A8"


def _hoja_cobranza(ws: Worksheet, req: BalanceAvionRequest) -> None:
    """Hoja 'cobranza' (26-ago-2026 · regla 28-ago): el estatus de cobro POR
    VUELO al frente — cuánto se debía cobrar (VENTA DEL AVIÓN), cuánto YA se
    cobró (prorrateado al avión) y cuánto falta; más el TOTAL COTIZACIÓN
    (c/extras), los depósitos REALES, la comisión que retuvo el banco y la
    cuenta que recibió cada parcialidad. Los mismos números del bloque
    STATUS DE COBROS del final de la hoja maestra, en formato legible."""
    n_cols = 13
    # Ancho (en caracteres) de la columna DETALLE DE COBROS: sirve para
    # estimar cuántos renglones envuelve cada parcialidad y no recortar.
    ancho_detalle = 46
    _title(ws, f"Cobranza — {req.matricula}".strip(" —"), 1, n_cols)
    periodo = f"Periodo: {req.periodo_desde or '—'} a {req.periodo_hasta or '—'}"
    ws.cell(row=2, column=1, value=periodo).font = Font(italic=True, size=10, color=MUTED)

    t = req.totales

    def _comision_banco(v: BalanceAvionVuelo) -> float | None:
        """Σ comisiones bancarias de las parcialidades (None si ninguna trae)."""
        montos = [c.comision_mxn for c in v.cobros if c.comision_mxn is not None]
        return round(sum(montos), 2) if montos else None

    def _suma(valores: list[float | None]) -> float | None:
        nums = [x for x in valores if x is not None]
        return round(sum(nums), 2) if nums else None

    # Totales de las columnas nuevas: los manda el API; si un API viejo no
    # los trae, se re-suman las filas SOLO para mostrar (nunca negocio).
    tot_cotizacion = t.total_cotizacion_mxn
    if tot_cotizacion is None:
        tot_cotizacion = _suma([v.total_cotizacion_mxn for v in req.vuelos])
    tot_cobrado_real = t.cobrado_real_mxn
    if tot_cobrado_real is None:
        tot_cobrado_real = _suma([v.cobrado_real_mxn for v in req.vuelos])
    tot_comision = t.comision_banco_mxn
    if tot_comision is None:
        tot_comision = _suma([_comision_banco(v) for v in req.vuelos])

    _header_row(ws, 4, ["TOTAL A COBRAR\nMXN (venta avión)", "COBRADO\nMXN",
                        "POR COBRAR\nMXN", "POR COBRAR\nUSD",
                        "TOTAL COTIZACIÓN\nMXN (c/extras)", "COBRADO REAL\nMXN",
                        "% COBRADO"])
    ws.row_dimensions[4].height = 30
    _num(ws, 5, 1, t.total_mxn, MONEY, bold=True)
    _num(ws, 5, 2, t.cobrado_mxn, MONEY, bold=True)
    pc = _num(ws, 5, 3, t.por_cobrar_mxn, MONEY, bold=True)
    if (t.por_cobrar_mxn or 0) > 0.005:
        pc.font = Font(bold=True, color=RED)
    _num(ws, 5, 4, t.por_cobrar_usd, MONEY)
    _num(ws, 5, 5, tot_cotizacion, MONEY, bold=True)
    _num(ws, 5, 6, tot_cobrado_real, MONEY, bold=True)
    # % COBRADO = COBRADO ÷ TOTAL A COBRAR: la MISMA base que POR COBRAR
    # (ambos al TC de venta). Nunca cobrado real ÷ total cotización: son
    # pesos a TCs distintos (el depósito trae su propio TC) y contradecían
    # el POR COBRAR de la misma fila (0 por cobrar y 97% cobrado).
    if t.total_mxn and t.total_mxn > 0:
        pct = (t.cobrado_mxn or 0) / t.total_mxn
    else:
        pct = None
    _num(ws, 5, 7, pct, "0.0%")
    for c in range(1, 8):
        ws.cell(row=5, column=c).border = _border

    headers = ["CLAVE", "FECHA", "RUTA", "ESTADO", "STATUS\nCOBRO",
               "TOTAL A COBRAR\nMXN (venta avión)", "COBRADO\nMXN",
               "POR COBRAR\nMXN", "POR COBRAR\nUSD",
               "TOTAL COTIZACIÓN\nMXN (c/extras)", "COBRADO REAL\nMXN",
               "COMISIÓN\nBANCO MXN",
               "DETALLE DE COBROS\n(fecha · monto · método · cuenta · comisión)"]
    _header_row(ws, 7, headers)
    ws.row_dimensions[7].height = 42
    row = 8
    for v in req.vuelos:
        cc = ws.cell(row=row, column=1, value=v.clave)
        avion_fill = _hex(v.avion_color)
        if avion_fill:
            cc.fill = PatternFill("solid", fgColor=avion_fill)
        ws.cell(row=row, column=2, value=_fecha(v.fecha))
        ws.cell(row=row, column=3, value=v.ruta)
        ec = ws.cell(row=row, column=4, value=v.estado)
        ec.font = Font(color=RED if v.estado == "CANCELADO" else MUTED, size=9)
        st = ws.cell(row=row, column=5, value=v.status_cobro or "—")
        if v.status_cobro == "Cobrado":
            st.fill = PatternFill("solid", fgColor="D1FAE5")
            st.font = Font(color="065F46", bold=True, size=9)
        elif v.status_cobro == "Parcial":
            st.fill = PatternFill("solid", fgColor=AMBER)
            st.font = Font(bold=True, size=9)
        elif v.status_cobro == "Pendiente":
            st.fill = PatternFill("solid", fgColor="FECACA")
            st.font = Font(color="991B1B", bold=True, size=9)
        _num(ws, row, 6, v.total_mxn)
        _num(ws, row, 7, v.cobrado_mxn)
        pcc = _num(ws, row, 8, v.por_cobrar_mxn)
        if (v.por_cobrar_mxn or 0) > 0.005:
            pcc.font = Font(color=RED)
        _num(ws, row, 9, v.por_cobrar_usd)
        _num(ws, row, 10, v.total_cotizacion_mxn)
        _num(ws, row, 11, v.cobrado_real_mxn)
        _num(ws, row, 12, _comision_banco(v))
        # Detalle: depósitos REALES (fecha · monto · método · cuenta ·
        # comisión) — las partes que no vienen se omiten.
        detalle = "\n".join(
            " · ".join(
                x
                for x in (
                    _fecha(c.fecha),
                    f"${c.monto_mxn:,.2f}" if c.monto_mxn is not None else None,
                    c.metodo or None,
                    c.cuenta or None,
                    (
                        f"comisión ${c.comision_mxn:,.2f}"
                        if c.comision_mxn is not None
                        else None
                    ),
                )
                if x
            )
            for c in v.cobros
        )
        dc = ws.cell(row=row, column=n_cols, value=detalle or None)
        dc.alignment = Alignment(wrap_text=True, vertical="top")
        # Alto de fila = renglones REALES del detalle: cada parcialidad
        # trae cuenta + comisión (~80 caracteres) y envuelve en 2 líneas
        # dentro de la columna; con la altura por cobro Excel recortaba.
        lineas = sum(
            max(1, math.ceil(len(linea) / ancho_detalle))
            for linea in detalle.split("\n")
        ) if detalle else 1
        if lineas > 1:
            ws.row_dimensions[row].height = 14 * lineas + 4
        for c in range(1, n_cols + 1):
            ws.cell(row=row, column=c).border = _border
        row += 1

    ws.cell(row=row, column=1, value="TOTALES").font = Font(bold=True)
    _num(ws, row, 6, t.total_mxn, MONEY, bold=True)
    _num(ws, row, 7, t.cobrado_mxn, MONEY, bold=True)
    _num(ws, row, 8, t.por_cobrar_mxn, MONEY, bold=True)
    _num(ws, row, 9, t.por_cobrar_usd, MONEY, bold=True)
    _num(ws, row, 10, tot_cotizacion, MONEY, bold=True)
    _num(ws, row, 11, tot_cobrado_real, MONEY, bold=True)
    _num(ws, row, 12, tot_comision, MONEY, bold=True)
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.border = _border
        cell.fill = PatternFill("solid", fgColor=LIGHT)
    row += 2

    for nota in (
        "TOTAL A COBRAR = venta del avión. COBRADO = parcialidades reales × "
        "(venta avión ÷ total cotización): la parte de los cobros que "
        "corresponde a TUAs/extras/pernocta es de VuelaTour. El detalle trae "
        "los depósitos reales y la comisión que retuvo el banco.",
        "TOTAL COTIZACIÓN = lo cobrado al cliente COMPLETO (con TUAs/extras/"
        "pernocta y su IVA); COBRADO REAL = depósitos tal cual entraron. "
        "% COBRADO = COBRADO ÷ TOTAL A COBRAR (misma base que POR COBRAR). "
        "COBRADO y POR COBRAR van al TC de venta: un depósito en pesos con "
        "su propio TC se convierte a USD con ese TC y se re-expresa al TC de "
        "venta, por lo que COBRADO puede diferir de COBRADO REAL sin que "
        "falte dinero.",
        "Filas COMPARTIDO y clientes INTERNOS van sin venta a propósito (la "
        "venta del compartido vive en el balance del avión principal; el "
        "interno no cobra). CANCELADO conserva sus cobros si los hubo.",
    ):
        ws.cell(row=row, column=1, value=nota).font = Font(
            color=MUTED, size=9, italic=True
        )
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
        row += 1

    for i, w in enumerate(
        [22, 11, 26, 12, 11, 16, 15, 15, 13, 16, 15, 14, ancho_detalle], start=1
    ):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A8"


def _hoja_combustible(ws: Worksheet, hoja: BalanceAvionHojaCombustible,
                      req: BalanceAvionRequest) -> None:
    """Pestaña 'combustible' (26-ago-2026): el gas del avión POR MES —
    ya no va por vuelo. Ledger con litros + resumen con $/L promedio."""
    _title(ws, f"Combustible — {req.matricula}".strip(" —"), 1, 6)
    periodo = f"Periodo: {req.periodo_desde or '—'} a {req.periodo_hasta or '—'}"
    ws.cell(row=2, column=1, value=periodo).font = Font(italic=True, size=10, color=MUTED)

    _header_row(ws, 4, ["TOTAL MXN", "LITROS", "$ X LITRO PROM", "TC PROMEDIO",
                        "TOTAL USD", "USD X HR"])
    _num(ws, 5, 1, hoja.total_mxn, MONEY, bold=True)
    _num(ws, 5, 2, hoja.litros_total, HORAS)
    _num(ws, 5, 3, hoja.precio_litro_prom, MONEY)
    _num(ws, 5, 4, req.totales.tc_promedio, TC)
    _num(ws, 5, 5, hoja.usd, MONEY, bold=True)
    _num(ws, 5, 6, hoja.usd_hr, MONEY)
    for c in range(1, 7):
        ws.cell(row=5, column=c).border = _border

    _header_row(ws, 7, ["FECHA", "DETALLE", "LITROS", "MONTO MXN",
                        "MONEDA ORIGINAL", "MONTO ORIGINAL"])
    row = 8
    if hoja.filas:
        for f in hoja.filas:
            ws.cell(row=row, column=1, value=_fecha(f.fecha)).border = _border
            ws.cell(row=row, column=2, value=f.detalle).border = _border
            _num(ws, row, 3, f.litros, HORAS).border = _border
            _num(ws, row, 4, f.monto_mxn).border = _border
            ws.cell(row=row, column=5, value=f.moneda_original).border = _border
            _num(ws, row, 6, f.monto_original).border = _border
            # Balance GENERAL: el DETALLE (lleva la matrícula al frente) se
            # tiñe con el color del avión — como su libro.
            avion_fill = _hex(f.avion_color)
            if avion_fill:
                ws.cell(row=row, column=2).fill = PatternFill(
                    "solid", fgColor=avion_fill
                )
            row += 1
    else:
        ws.cell(
            row=row, column=1,
            value="Sin cargas de combustible en el periodo.",
        ).font = Font(color=MUTED, italic=True)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1

    row += 1
    ws.cell(
        row=row, column=1,
        value="El combustible se controla POR AVIÓN y POR MES (fecha del "
        "gasto, con o sin vuelo) — mismo criterio que el reparto a socios. "
        "Resta una sola vez en la hoja 'balance'.",
    ).font = Font(color=MUTED, size=9, italic=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)

    for i, w in enumerate([14, 52, 10, 15, 16, 15], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A8"


def _hoja_balance(ws: Worksheet, req: BalanceAvionRequest) -> None:
    _title(ws, f"Balance — {req.matricula}".strip(" —"), 1, 3)
    periodo = f"Periodo: {req.periodo_desde or '—'} a {req.periodo_hasta or '—'} (todo en USD)"
    ws.cell(row=2, column=1, value=periodo).font = Font(italic=True, size=10, color=MUTED)
    _bloque_balance(ws, req, 4)
    for i, w in enumerate([46, 14, 16], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _bloque_balance(ws: Worksheet, req: BalanceAvionRequest, row: int) -> int:
    """Bloque utilidad + reparto de socios de UN avión, empezando en `row`.
    Devuelve la siguiente fila libre (el balance GENERAL apila un bloque por
    avión — los socios son POR avión, no hay un reparto de flota)."""
    b = req.balance
    filas: list[tuple[str, float | None, bool]] = [
        ("UTILIDAD ANTES DE GASTOS USD", b.utilidad_antes_usd, False),
        ("(−) COMBUSTIBLE DEL MES USD", b.combustible_usd, False),
        ("(−) GASTOS INDIRECTOS USD", b.gastos_indirectos_usd, False),
        ("(−) OTROS GASTOS USD", b.otros_usd, False),
        ("(−) PERMISOS USD", b.permisos_usd, False),
        ("UTILIDAD DESPUÉS DE GASTOS USD", b.utilidad_despues_usd, True),
        ("(−) PENDIENTE DE PAGO (COBRANZA PENDIENTE) USD", b.por_cobrar_usd, False),
        ("UTILIDAD COBRADA USD", b.utilidad_cobrada_usd, True),
    ]
    for label, val, bold in filas:
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = Font(bold=bold, color=NAVY if bold else MUTED)
        lc.border = _border
        color = None
        if label.startswith("UTILIDAD COBRADA") and val is not None:
            color = GREEN if val >= 0 else RED
        cell = _num(ws, row, 2, val, MONEY, bold=bold, color=color or ("000000" if bold else MUTED))
        cell.border = _border
        row += 1

    row += 1
    _title(ws, "Reparto a socios (utilidad COBRADA × % vigente)", row, 3, size=11)
    row += 1
    _header_row(ws, row, ["SOCIO", "PORCENTAJE", "MONTO USD"])
    row += 1
    if b.socios:
        for s in b.socios:
            ws.cell(row=row, column=1, value=s.nombre).border = _border
            pc = _num(ws, row, 2, s.porcentaje, '0.00"%"')
            pc.border = _border
            mc = _num(ws, row, 3, s.monto_usd, MONEY, bold=True)
            mc.border = _border
            row += 1
    else:
        ws.cell(
            row=row, column=1, value="Sin socios configurados para este avión (ver pendientes)."
        ).font = Font(color=RED, italic=True)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        row += 1
    return row


def _hoja_pendientes(ws: Worksheet, req: BalanceAvionRequest) -> None:
    _title(ws, f"Pendientes de captura — {req.matricula}".strip(" —"), 1, 2)
    ws.cell(
        row=2, column=1,
        value="Genera de nuevo el balance después de capturar; esta hoja debe quedar vacía.",
    ).font = Font(italic=True, size=10, color=MUTED)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)

    row = 4
    if req.pendientes:
        for i, texto in enumerate(req.pendientes, start=1):
            nc = ws.cell(row=row, column=1, value=i)
            nc.font = Font(bold=True, color=RED)
            nc.alignment = Alignment(horizontal="right", vertical="top")
            ws.cell(row=row, column=2, value=texto).alignment = Alignment(
                wrap_text=True, vertical="top"
            )
            row += 1
    else:
        c = ws.cell(
            row=row, column=2, value="Sin pendientes — la captura del periodo está completa."
        )
        c.font = Font(bold=True, color=GREEN)

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 95
    ws.freeze_panes = "A4"


# Definiciones de las hojas de gastos (regla del cliente, 28-ago-2026).
_NOTA_INDIRECTOS = (
    "GASTOS INDIRECTOS = gastos del avión que no se pudieron ligar a un vuelo "
    "(capturados con la aeronave, sin vuelo): cualquier categoría salvo "
    "combustible (hoja 'combustible') y permisos (hoja 'permisos')."
)
_NOTA_PERMISOS = "PERMISOS = todo lo de AFAC (permisos y provisiones del avión)."


def _nota_otros_gastos(general: bool) -> str:
    parte = (
        "solo la parte asignada a cada avión (fila con su color)"
        if general
        else "solo la parte de este avión"
    )
    return (
        "OTROS GASTOS = gastos administrativos de la empresa (nómina, IMSS, "
        f"pensión, fijos…) repartidos a mano entre aviones — aquí {parte}. "
        "El TUA pagado NO va aquí (regla 28-ago-2026): queda solo como nota "
        "en OPERACIONES y vive en 'otros movimientos' del Balance general."
    )


def render_balance_avion_xlsx(req: BalanceAvionRequest) -> bytes:
    wb = Workbook()
    _hoja_maestra(wb.active, req)
    _hoja_cobranza(wb.create_sheet("cobranza"), req)
    _hoja_combustible(wb.create_sheet("combustible"), req.combustible, req)
    _hoja_gastos(wb.create_sheet("gastos indirectos"), "Gastos indirectos",
                 req.gastos_indirectos, req, nota=_NOTA_INDIRECTOS)
    _hoja_gastos(
        wb.create_sheet("otros gastos"), "Otros gastos", req.otros_gastos, req,
        nota=_nota_otros_gastos(general=False),
    )
    _hoja_gastos(wb.create_sheet("permisos"), "Permisos", req.permisos, req,
                 nota=_NOTA_PERMISOS)
    _hoja_balance(wb.create_sheet("balance"), req)
    _hoja_pendientes(wb.create_sheet("pendientes de captura"), req)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _fila_resumen(ws: Worksheet, row: int, f: BalanceGeneralResumenFila,
                  *, bold: bool = False) -> None:
    c = ws.cell(row=row, column=1, value=f.matricula)
    if bold:
        c.font = Font(bold=True)
    # La celda de la matrícula teñida con el color del avión = la LEYENDA del
    # libro (tabla "Color calendario" del equipo).
    swatch = _hex(f.color)
    if swatch:
        c.fill = PatternFill("solid", fgColor=swatch)
    _num(ws, row, 2, f.vuelos, "0", **({"bold": True} if bold else {}))
    _num(ws, row, 3, f.horas, HORAS, **({"bold": True} if bold else {}))
    _num(ws, row, 4, f.horas_cobradas, HORAS, **({"bold": True} if bold else {}))
    _num(ws, row, 5, f.venta_mxn, MONEY, **({"bold": True} if bold else {}))
    _num(ws, row, 6, f.costo_mxn, MONEY, **({"bold": True} if bold else {}))
    _num(ws, row, 7, f.combustible_mxn, MONEY, **({"bold": True} if bold else {}))
    _num(ws, row, 8, f.comisiones_mxn, MONEY, **({"bold": True} if bold else {}))
    _num(ws, row, 9, f.ganancia_mxn, MONEY, **({"bold": True} if bold else {}))
    _num(ws, row, 10, f.cobrado_mxn, MONEY, **({"bold": True} if bold else {}))
    _num(ws, row, 11, f.por_cobrar_mxn, MONEY, **({"bold": True} if bold else {}))
    _num(ws, row, 12, f.pendientes, "0", **({"bold": True} if bold else {}))


def _hoja_resumen_general(ws: Worksheet, req: BalanceGeneralRequest) -> None:
    """Índice de la flota: una fila por avión (los números vienen del API —
    son los TOTALES del libro de cada avión, jamás se recalculan aquí)."""
    ws.title = "RESUMEN flota"
    _title(
        ws,
        f"Balance general de flota · {req.periodo_desde or ''} a {req.periodo_hasta or ''}",
        1,
        12,
    )
    _header_row(ws, 3, [
        "AVIÓN", "VUELOS", "HORAS\nVOLADAS", "HORAS\nCOBRADAS", "VENTA\nMXN",
        "COSTO TOTAL\nMXN", "COMBUSTIBLE\nMXN", "COMISIONES\nVENDEDOR MXN",
        "GANANCIA\nMXN", "COBRADO\nMXN", "POR COBRAR\nMXN", "PENDIENTES",
    ])
    row = 4
    for f in req.resumen:
        _fila_resumen(ws, row, f)
        row += 1
    if req.resumen_totales is not None:
        _fila_resumen(ws, row, req.resumen_totales, bold=True)
        row += 1
    row += 1
    for nota in (
        "VENTA = tiempo de vuelo + ajuste + IVA proporcional (sin TUAs/extras/"
        "pernocta: ver 'otros movimientos'). GANANCIA = VENTA − COSTO TOTAL − "
        "COMBUSTIBLE − COMISIONES; los TUAs pagados no restan a ningún avión. "
        "La GANANCIA va antes de otros/indirectos/permisos — la utilidad "
        "final por avión está en la hoja 'balance'.",
        "Vuelos multi-avión: los costos de columnas de la fila COMPARTIDO "
        "suman al COSTO de su avión pero su ganancia va vacía (prorrateo "
        "pendiente) — en esos meses el cruce de columnas difiere por ese "
        "monto.",
        "El detalle está en las hojas siguientes: los datos de TODOS los "
        "aviones juntos, en el mismo orden del libro individual — cada fila "
        "se identifica por su CLAVE y el color de su avión. El libro "
        "individual de cada avión se descarga desde su ficha.",
    ):
        ws.cell(row=row, column=1, value=nota).font = Font(
            color=MUTED, size=9, italic=True
        )
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
        row += 1
    for i in range(1, 13):
        ws.column_dimensions[get_column_letter(i)].width = 10 if i == 1 else 14
    ws.freeze_panes = "A4"


# Estados de vuelo que NO se anotan junto a la clave en 'otros movimientos'
# (los normales); cualquier otro (RESERVA, CANCELADO…) se marca.
_ESTADOS_NORMALES = ("CONFIRMADO", "EN_VUELO", "COMPLETADO")


def _hoja_otros_movimientos(ws: Worksheet, hoja: BalanceHojaOtrosMovimientos) -> None:
    """Pestaña "Otros movimientos" (28-ago, réplica de la hoja manual
    "dinero otros ingresos"): conceptos cobrados al cliente vs pagados, por
    clave de vuelo (celda teñida con el color del avión), más la sección de
    movimientos SIN avión/SIN vuelo. Remanente negativo en ROJO. Incluye
    TODOS los estados del periodo (igual que la hoja maestra): la clave
    lleva ' · ESTADO' cuando no es un estado normal y va en ROJO si es
    CANCELADO. Los montos vienen YA calculados del API — aquí solo se
    pinta."""
    ws.cell(row=1, column=1, value="OTROS MOVIMIENTOS VUELATOUR").font = Font(
        bold=True, size=12, color=NAVY
    )
    ws.cell(
        row=2, column=1,
        value="Ingreso de VuelaTour (no del avión): TUAs, extras y viáticos de "
        "pernocta cobrados al cliente (con su IVA) vs lo pagado, por concepto; "
        "el egreso solo se aparea cuando el mapeo es directo (TUAs, hotel de "
        "pernocta, comisión bancaria) — el resto se lee por clave. Una fila "
        "puede traer solo ingreso o solo egreso. Todo en MXN. Incluye todos "
        "los estados del periodo (igual que la hoja maestra); los cancelados "
        "se marcan (clave · CANCELADO en rojo) y los demás estados no "
        "normales llevan su estado junto a la clave.",
    ).font = Font(italic=True, size=9, color=MUTED)
    headers = [
        "clave", "fecha\nvuelo", "concepto", "egreso", "fecha", "concepto",
        "ingreso", "fecha", "remanente", "factura\nvuelatour",
    ]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=9)
        c.fill = PatternFill("solid", fgColor=BRAND)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
        ws.column_dimensions[get_column_letter(i)].width = 15
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["F"].width = 32
    row = 4
    tot_e = tot_i = 0.0

    def pinta(f: BalanceOtroMovimientoFila) -> None:
        nonlocal row, tot_e, tot_i
        # Estado del vuelo (API nuevo): se anota junto a la clave cuando NO
        # es normal; CANCELADO además en rojo (igual que la hoja maestra).
        estado = (f.estado or "").strip().upper()
        clave = f.clave or ""
        if estado and estado not in _ESTADOS_NORMALES:
            clave = f"{clave} · {estado}" if clave else estado
        vals = [
            clave, _fecha(f.fecha_vuelo), f.concepto_egreso, f.egreso_mxn,
            _fecha(f.fecha_egreso), f.concepto_ingreso, f.ingreso_mxn,
            _fecha(f.fecha_ingreso), f.remanente_mxn, f.factura,
        ]
        for i, v in enumerate(vals, start=1):
            cell = ws.cell(row=row, column=i, value=v if v != "" else None)
            cell.border = _border
            if isinstance(v, (int, float)):
                cell.number_format = MONEY
        swatch = _hex(f.avion_color)
        if swatch:
            ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=swatch)
        if estado == "CANCELADO":
            ws.cell(row=row, column=1).font = Font(color=RED)
        if isinstance(f.remanente_mxn, (int, float)) and f.remanente_mxn < 0:
            ws.cell(row=row, column=9).font = Font(color=RED)
        if isinstance(f.egreso_mxn, (int, float)):
            tot_e += f.egreso_mxn
        if isinstance(f.ingreso_mxn, (int, float)):
            tot_i += f.ingreso_mxn
        row += 1

    for f in hoja.filas:
        pinta(f)
    if not hoja.filas and not hoja.filas_sueltas:
        ws.cell(
            row=row, column=1, value="Sin otros movimientos en el periodo."
        ).font = Font(color=MUTED, italic=True)
        row += 1
    if hoja.filas_sueltas:
        row += 1
        t = ws.cell(row=row, column=1, value="MOVIMIENTOS SIN AVIÓN / SIN VUELO")
        t.font = Font(bold=True, size=10, color=NAVY)
        t.fill = PatternFill("solid", fgColor=LIGHT)
        for i in range(2, 11):
            ws.cell(row=row, column=i).fill = PatternFill("solid", fgColor=LIGHT)
        row += 1
        for f in hoja.filas_sueltas:
            pinta(f)
    ws.cell(row=row, column=1, value="TOTALES").font = Font(bold=True)
    rem = tot_i - tot_e
    for col, val in ((4, round(tot_e, 2)), (7, round(tot_i, 2)), (9, round(rem, 2))):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = Font(bold=True, color=RED if col == 9 and rem < 0 else "000000")
        cell.number_format = MONEY
        cell.fill = PatternFill("solid", fgColor=LIGHT)
    ws.freeze_panes = "A4"


def render_balance_general_xlsx(req: BalanceGeneralRequest) -> bytes:
    """Balance GENERAL (regla del cliente, 18-ago): la MISMA estructura de
    hojas que el libro individual pero con los datos de TODOS los aviones
    JUNTOS — 1 reporte de horas, 1 combustible (mensual), 1 gastos
    indirectos, 1 otros gastos (parte repartida a cada avión; sin TUAs),
    1 permisos, 1 otros movimientos (TUAs/extras/pernocta cobrados y
    pagados: dinero de VuelaTour), 1 balance (bloques por avión: los socios
    son por avión) y 1 pendientes. Cada
    fila se identifica por su clave y el COLOR del avión
    (aeronave.color_calendario, editable en el apartado del avión)."""
    wb = Workbook()
    _hoja_resumen_general(wb.active, req)
    cons = req.consolidado
    if cons is not None:
        # La hoja maestra se titula sola ("reporte horas FLOTA").
        _hoja_maestra(wb.create_sheet(), cons)
        _hoja_cobranza(wb.create_sheet("cobranza"), cons)
        _hoja_combustible(wb.create_sheet("combustible"), cons.combustible, cons)
        _hoja_gastos(
            wb.create_sheet("gastos indirectos"),
            "Gastos indirectos", cons.gastos_indirectos, cons,
            nota=_NOTA_INDIRECTOS,
        )
        _hoja_gastos(
            wb.create_sheet("otros gastos"), "Otros gastos",
            cons.otros_gastos, cons,
            nota=_nota_otros_gastos(general=True),
        )
        _hoja_gastos(wb.create_sheet("permisos"), "Permisos", cons.permisos, cons,
                     nota=_NOTA_PERMISOS)

        # Pestaña "Otros movimientos" (28-ago): solo si el API la manda.
        if cons.otros_movimientos is not None:
            _hoja_otros_movimientos(
                wb.create_sheet("otros movimientos"), cons.otros_movimientos
            )

        # Hoja balance: un BLOQUE por avión (título teñido con su color).
        ws_b = wb.create_sheet("balance")
        _title(ws_b, "Balance por avión", 1, 3)
        ws_b.cell(
            row=2, column=1,
            value=f"Periodo: {req.periodo_desde or '—'} a "
            f"{req.periodo_hasta or '—'} (todo en USD) · los socios son POR avión",
        ).font = Font(italic=True, size=10, color=MUTED)
        row = 4
        for avion in req.aviones:
            tc = ws_b.cell(row=row, column=1, value=avion.matricula or "—")
            tc.font = Font(bold=True, size=12, color=NAVY)
            # Solo la celda de la matrícula lleva el color (como su libro).
            swatch = _hex(avion.avion_color)
            if swatch:
                tc.fill = PatternFill("solid", fgColor=swatch)
            row = _bloque_balance(ws_b, avion, row + 1) + 2
        for i, w in enumerate([46, 14, 16], start=1):
            ws_b.column_dimensions[get_column_letter(i)].width = w

        _hoja_pendientes(wb.create_sheet("pendientes de captura"), cons)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
