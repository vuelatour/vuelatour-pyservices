"""Libro "Dinero <periodo>" (openpyxl): réplica del control manual del equipo
("Dinero Junio COMPLETO.xlsx"), mismas hojas y mismo ORDEN de columnas:

  1. dinero-vlos    — una fila por vuelo (venta, costos proveedor, comisión,
                      status de cobros/pagos). Fila coloreada con el color del
                      avión (el de su calendario) y la CLAVE lleva nota con la
                      matrícula. Las columnas SIN regla en el sistema todavía
                      (costo proveedor, comisiones, pagos) van VACÍAS
                      conservando su lugar. VENTA AVIÓN (regla 28-ago-2026)
                      = tiempo + ajuste + IVA proporcional; el TOTAL FACTURA
                      AL CLIENTE (con TUAs/extras/pernocta) va junto al
                      STATUS DE COBROS como informativo.
  2. Otros ingresos — TUAs/extras/pernocta (+ su IVA) por vuelo: ingreso de
                      VuelaTour, no del avión (ingreso vs egreso).
  3. otros gastos   — gastos del mes sin vuelo, con acumulado.
  4. utilidades     — resumen del periodo (lo computable hoy).

El API manda TODO calculado; aquí solo se pinta (regla del repo).
"""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.schemas.reportes import DineroVueloFila, DineroXlsxRequest

NAVY = "102A43"
LIGHT = "EEF2F7"
MONEY = '"$"#,##0.00'
HORAS = "0.0"
TC = "0.0000"

_thin = Side(style="thin", color="D5DBE3")
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_CANCUN = ZoneInfo("America/Cancun")


def _fecha(s: str | None) -> str:
    if not s:
        return ""
    try:
        dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
        if dt.tzinfo is not None:
            dt = dt.astimezone(_CANCUN)
        return dt.strftime("%d/%m/%Y")
    except ValueError:
        return s


def _hex(color: str | None) -> str | None:
    """#RRGGBB → RRGGBB aclarado NO: se usa tal cual (libro del equipo usa
    colores saturados por avión)."""
    if not color:
        return None
    c = color.lstrip("#").strip()
    return c.upper() if len(c) == 6 else None


# Columnas de la hoja dinero-vlos en el ORDEN del libro (A..BE). Cada entrada:
# (grupo, encabezado, attr de DineroVueloFila o None, formato o None).
# attr None + fmt None = columna sin regla todavía (celda vacía) o especial.
_COLS: list[tuple[str, str, str | None, str | None]] = [
    ("", "CLAVE", "clave", None),
    ("", "FECHA", "fecha", None),
    ("", "RUTA", "ruta", None),
    ("VENTA", "TIEMPO\nCALZOS HOBS\ncobrado", "tiempo", HORAS),
    ("VENTA", "VENTA HR\nSIN IVA\ndllrs", "venta_hr_usd", MONEY),
    ("VENTA", "VENTA HR\nSIN IVA\npesos", "venta_hr_mxn", MONEY),
    ("VENTA", "IVA X\nHR dllrs", "iva_hr_usd", MONEY),
    ("VENTA", "VENTA HR\nMAS IVA\ndllrs", "venta_hr_masiva_usd", MONEY),
    # VENTA AVIÓN (regla 28-ago-2026) = tiempo + ajuste + IVA proporcional;
    # los TUAs/extras/pernocta cobrados van en la hoja "Otros ingresos".
    ("VENTA", "VENTA AVIÓN\n(tiempo+ajuste+IVA)\ndllrs", "total_cobrado_usd", MONEY),
    ("VENTA", "IVA VENTA\nAVIÓN dllrs", "iva_total_usd", MONEY),
    ("VENTA", "TIPO\nCAMBIO", "tc_venta", TC),
    ("VENTA", "VENTA AVIÓN\n(tiempo+ajuste+IVA)\npesos", "total_cobrado_mxn", MONEY),
    ("VENTA", "IVA VENTA\nAVIÓN pesos", "iva_total_mxn", MONEY),
    ("VENTA", "VENTA AVIÓN\nS/IVA pesos", "total_siva_mxn", MONEY),
    ("COSTO PROVEEDOR", "TIEMPO\nCALZOS HOBS\npagado", None, HORAS),
    ("COSTO PROVEEDOR", "COSTO X\nHORA dllrs", None, MONEY),
    ("COSTO PROVEEDOR", "IVA\nX HR dllrs", None, MONEY),
    ("COSTO PROVEEDOR", "COSTO HR\nMAS IVA\ndllrs", None, MONEY),
    ("COSTO PROVEEDOR", "TOTAL PARA\nPROVEEDOR\ndllrs", None, MONEY),
    ("COSTO PROVEEDOR", "IVA TOTAL\nPAGADO dllrs", None, MONEY),
    ("COSTO PROVEEDOR", "TIPO\nCAMBIO", None, TC),
    ("COSTO PROVEEDOR", "TOTAL PARA\nPROVEEDOR\npesos", None, MONEY),
    ("COSTO PROVEEDOR", "IVA TOTAL\nPAGADO pesos", None, MONEY),
    ("COSTO PROVEEDOR", "TOTAL PAGADO\nS/IVA pesos", None, MONEY),
    ("", "REMANENTE\nVENTA−COMPRA\npesos", None, MONEY),
    ("COMISIÓN", "COMISIONISTA", None, None),
    ("COMISIÓN", "STATUS", None, None),
    ("COMISIÓN", "COMISION\nX HR dllrs", None, MONEY),
    ("COMISIÓN", "TOTAL\nCOMISION", None, MONEY),
    ("COMISIÓN", "TIPO\nCAMBIO", None, TC),
    ("COMISIÓN", "TOTAL\nCOMISION\npesos", None, MONEY),
    ("COMISIÓN", "PORCENTAJE\nCOMISION", None, "0.00%"),
    ("", "OTROS\nGASTOS", None, MONEY),
    ("", "GANANCIA\nDESPUES\nCOMISIONES\npesos", None, MONEY),
    ("", "GANANCIA\nX HR pesos", None, MONEY),
    ("", "PORCENTAJE\nGANANCIA", None, "0.00%"),
    # Total COMPLETO facturado al cliente (con TUAs/extras/pernocta + IVA):
    # informativo junto a los cobros — contra esto se cuadran los depósitos.
    ("STATUS DE COBROS", "TOTAL FACTURA\nAL CLIENTE\npesos", "total_cliente_mxn", MONEY),
    ("STATUS DE COBROS", "STATUS", "status_cobro", None),
    ("STATUS DE COBROS", "COBRO 1\nFECHA", None, None),
    ("STATUS DE COBROS", "COBRO 1\nCANTIDAD", None, MONEY),
    ("STATUS DE COBROS", "COBRO 2\nFECHA", None, None),
    ("STATUS DE COBROS", "COBRO 2\nCANTIDAD", None, MONEY),
    ("STATUS DE COBROS", "COBRO 3\nFECHA", None, None),
    ("STATUS DE COBROS", "COBRO 3\nCANTIDAD", None, MONEY),
    ("STATUS DE COBROS", "COBRO 4\nFECHA", None, None),
    ("STATUS DE COBROS", "COBRO 4\nCANTIDAD", None, MONEY),
    ("STATUS DE COBROS", "TOTAL", "total_cobros_mxn", MONEY),
    ("STATUS DE COBROS", "ME DEBEN", "me_deben_mxn", MONEY),
    ("STATUS DE PAGOS", "STATUS", None, None),
    ("STATUS DE PAGOS", "PAGO 1\nFECHA", None, None),
    ("STATUS DE PAGOS", "PAGO 1\nCANTIDAD", None, MONEY),
    ("STATUS DE PAGOS", "PAGO 2\nFECHA", None, None),
    ("STATUS DE PAGOS", "PAGO 2\nCANTIDAD", None, MONEY),
    ("STATUS DE PAGOS", "PAGO 3\nFECHA", None, None),
    ("STATUS DE PAGOS", "PAGO 3\nCANTIDAD", None, MONEY),
    ("STATUS DE PAGOS", "DEBO", None, MONEY),
    ("", "FACTURA\nVUELATOUR", "factura_vuelatour", None),
    ("", "FACTURA\nPROVEEDOR", None, None),
]
_COBRO1_COL = next(i for i, c in enumerate(_COLS, start=1) if c[1] == "COBRO 1\nFECHA")
_GROUP_FILLS = {
    "VENTA": "DDEBF7",
    "COSTO PROVEEDOR": "FCE4D6",
    "COMISIÓN": "E2EFDA",
    "STATUS DE COBROS": "FFF2CC",
    "STATUS DE PAGOS": "EDEDED",
}
# Suma de columnas en la fila TOTALES (atributos numéricos con regla).
_TOTAL_ATTRS = {
    "tiempo",
    "total_cobrado_usd",
    "iva_total_usd",
    "total_cobrado_mxn",
    "iva_total_mxn",
    "total_siva_mxn",
    "total_cliente_mxn",
    "total_cobros_mxn",
    "me_deben_mxn",
}


def _cobros_a_4(v: DineroVueloFila):
    cobros = v.cobros
    if len(cobros) <= 4:
        return cobros
    resto = cobros[3:]
    montos = [c.monto_mxn for c in resto if c.monto_mxn is not None]
    agg = round(sum(montos), 2) if montos else None
    ultima = next((c.fecha for c in reversed(resto) if c.fecha), None)
    from app.schemas.reportes import DineroCobroPago

    etiqueta = f"{_fecha(ultima)} (+{len(resto)})" if ultima else f"(+{len(resto)})"
    return [*cobros[:3], DineroCobroPago(fecha=etiqueta, monto_mxn=agg)]


def _hoja_vuelos(ws: Worksheet, req: DineroXlsxRequest) -> None:
    ws.title = "dinero-vlos"
    n = len(_COLS)
    t = ws.cell(row=1, column=1, value="RELACION DE DINERO INGRESADO X VUELOS VUELATOUR")
    t.font = Font(bold=True, size=12, color=NAVY)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
    sub = ws.cell(
        row=2,
        column=1,
        value=f"Periodo {_fecha(req.periodo_desde)} — {_fecha(req.periodo_hasta)} · generado {_fecha(req.generado)} (hora Cancún)",
    )
    sub.font = Font(italic=True, size=9, color="5B6470")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=12)

    # Encabezado de 2 filas: grupo (merged) + columna.
    col = 1
    while col <= n:
        grupo = _COLS[col - 1][0]
        fin = col
        while fin < n and _COLS[fin][0] == grupo:
            fin += 1
        if grupo:
            c = ws.cell(row=3, column=col, value=grupo)
            c.font = Font(bold=True, color="FFFFFF", size=9)
            c.fill = PatternFill("solid", fgColor=NAVY)
            c.alignment = Alignment(horizontal="center", vertical="center")
            if fin > col:
                ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=fin)
        col = fin + 1
    for i, (grupo, header, _attr, _fmt) in enumerate(_COLS, start=1):
        c = ws.cell(row=4, column=i, value=header)
        c.font = Font(bold=True, color=NAVY, size=7)
        c.fill = PatternFill("solid", fgColor=_GROUP_FILLS.get(grupo, LIGHT))
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
        ws.column_dimensions[get_column_letter(i)].width = 11
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["C"].width = 20
    ws.row_dimensions[4].height = 34

    totales: dict[str, float] = {}
    row = 5
    for v in req.vuelos:
        fill_hex = _hex(v.color)
        fila_fill = PatternFill("solid", fgColor=fill_hex) if fill_hex else None
        cobros = _cobros_a_4(v)
        for i, (_grupo, _header, attr, fmt) in enumerate(_COLS, start=1):
            val = None
            if attr is not None:
                val = getattr(v, attr)
                if attr == "fecha":
                    val = _fecha(val)
            elif _COBRO1_COL <= i < _COBRO1_COL + 8:
                idx = (i - _COBRO1_COL) // 2
                cobro = cobros[idx] if idx < len(cobros) else None
                if (i - _COBRO1_COL) % 2 == 0:
                    val = _fecha(cobro.fecha) if cobro and cobro.fecha else None
                else:
                    val = cobro.monto_mxn if cobro else None
            cell = ws.cell(row=row, column=i, value=val)
            cell.border = _border
            if fmt and isinstance(val, (int, float)):
                cell.number_format = fmt
            if fila_fill:
                cell.fill = fila_fill
            if attr == "clave":
                cell.font = Font(size=9)
                if v.matricula:
                    nota = Comment(
                        f"{v.matricula}", "VuelaTour", width=140, height=40
                    )
                    cell.comment = nota
            if attr is not None and isinstance(val, (int, float)) and attr in _TOTAL_ATTRS:
                totales[attr] = totales.get(attr, 0.0) + float(val)
        row += 1

    tot = ws.cell(row=row, column=1, value="TOTALES")
    tot.font = Font(bold=True)
    for i, (_grupo, _header, attr, fmt) in enumerate(_COLS, start=1):
        cell = ws.cell(row=row, column=i)
        cell.border = _border
        cell.fill = PatternFill("solid", fgColor=LIGHT)
        if attr in _TOTAL_ATTRS and attr in totales:
            cell.value = round(totales[attr], 2)
            cell.font = Font(bold=True)
            if fmt:
                cell.number_format = fmt

    # Leyenda de colores por avión (como el libro: "COLORES VIGENTES…").
    lrow = row + 2
    ws.cell(row=lrow, column=1, value="COLORES POR AVIÓN").font = Font(
        bold=True, size=9, color=NAVY
    )
    for item in req.leyenda_colores:
        lrow += 1
        mat = str(item.get("matricula", ""))
        hexc = _hex(item.get("color"))
        c = ws.cell(row=lrow, column=1, value=mat)
        c.border = _border
        if hexc:
            c.fill = PatternFill("solid", fgColor=hexc)
        ws.cell(row=lrow, column=2, value=str(item.get("modelo", ""))).border = _border

    nrow = lrow + 2
    ws.cell(
        row=nrow,
        column=1,
        value="VENTA AVIÓN = tiempo de vuelo + ajuste + IVA proporcional (regla "
        "28-ago-2026). Los TUAs, extras y viáticos de pernocta cobrados NO son "
        "venta del avión: son ingreso de VuelaTour (hoja 'Otros ingresos'). "
        "TOTAL FACTURA AL CLIENTE = lo cobrado completo (con TUAs/extras/"
        "pernocta y su IVA) — contra eso cuadran los cobros.",
    ).font = Font(italic=True, size=9, color="5B6470")
    ws.merge_cells(start_row=nrow, start_column=1, end_row=nrow, end_column=14)

    ws.freeze_panes = ws.cell(row=5, column=4)


def _hoja_otros_ingresos(ws: Worksheet, req: DineroXlsxRequest) -> None:
    ws.title = "Otros ingresos"
    ws.cell(row=1, column=1, value="RELACION DE OTROS INGRESOS VUELATOUR").font = Font(
        bold=True, size=12, color=NAVY
    )
    headers = [
        "clave", "fecha\nvuelo", "concepto", "egreso", "fecha", "concepto",
        "ingreso", "fecha", "remanente", "factura\nvuelatour",
    ]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=9)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
        ws.column_dimensions[get_column_letter(i)].width = 15
    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["F"].width = 26
    row = 4
    tot_e = tot_i = 0.0
    for f in req.otros_ingresos:
        vals = [
            f.clave, _fecha(f.fecha_vuelo), f.concepto_egreso, f.egreso_mxn,
            _fecha(f.fecha_egreso), f.concepto_ingreso, f.ingreso_mxn,
            _fecha(f.fecha_ingreso), f.remanente_mxn, f.factura,
        ]
        for i, v in enumerate(vals, start=1):
            cell = ws.cell(row=row, column=i, value=v if v != "" else None)
            cell.border = _border
            if isinstance(v, (int, float)):
                cell.number_format = MONEY
        if isinstance(f.egreso_mxn, (int, float)):
            tot_e += f.egreso_mxn
        if isinstance(f.ingreso_mxn, (int, float)):
            tot_i += f.ingreso_mxn
        row += 1
    ws.cell(row=row, column=1, value="TOTALES").font = Font(bold=True)
    for col, val in ((4, tot_e), (7, tot_i), (9, tot_i - tot_e)):
        cell = ws.cell(row=row, column=col, value=round(val, 2))
        cell.font = Font(bold=True)
        cell.number_format = MONEY
        cell.fill = PatternFill("solid", fgColor=LIGHT)
    ws.freeze_panes = "A4"


def _hoja_otros_gastos(ws: Worksheet, req: DineroXlsxRequest) -> None:
    ws.title = "otros gastos"
    ws.cell(row=1, column=1, value="RELACION DE OTROS GASTOS MENSUALES").font = Font(
        bold=True, size=12, color=NAVY
    )
    ws.cell(
        row=2, column=1,
        value="El combustible ya no va aquí: tiene su pestaña 'Combustible' "
        "(26-ago-2026).",
    ).font = Font(italic=True, size=9, color="5B6470")
    for i, h in enumerate(["fecha", "concepto", "monto", "acumulado"], start=1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=9)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center")
        c.border = _border
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 44
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    row = 4
    for g in req.otros_gastos:
        vals = [_fecha(g.fecha), g.concepto, g.monto_mxn, g.acumulado_mxn]
        for i, v in enumerate(vals, start=1):
            cell = ws.cell(row=row, column=i, value=v if v != "" else None)
            cell.border = _border
            if isinstance(v, (int, float)):
                cell.number_format = MONEY
        row += 1
    ws.freeze_panes = "A4"


def _hoja_combustible(ws: Worksheet, req: DineroXlsxRequest) -> None:
    """Pestaña 'Combustible' (26-ago-2026): el gas del MES por avión (con o
    sin vuelo, por fecha del gasto) — mismo criterio que el reparto. Las
    cargas SIN avión van marcadas: hay que asignarles aeronave."""
    ws.title = "Combustible"
    ws.cell(
        row=1, column=1, value="GASTO DE COMBUSTIBLE DEL MES (POR AVIÓN)"
    ).font = Font(bold=True, size=12, color=NAVY)
    for i, h in enumerate(
        ["fecha", "avión", "concepto", "litros", "monto", "acumulado"], start=1
    ):
        c = ws.cell(row=3, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=9)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center")
        c.border = _border
    for col, w in zip("ABCDEF", [12, 10, 44, 10, 14, 14], strict=True):
        ws.column_dimensions[col].width = w
    row = 4
    for g in req.combustible:
        vals = [_fecha(g.fecha), g.matricula, g.concepto, g.litros,
                g.monto_mxn, g.acumulado_mxn]
        for i, v in enumerate(vals, start=1):
            cell = ws.cell(row=row, column=i, value=v if v != "" else None)
            cell.border = _border
            if isinstance(v, (int, float)):
                cell.number_format = MONEY if i >= 5 else "0.0"
        # La celda del avión se tiñe con su color (leyenda del libro).
        swatch = _hex(g.avion_color)
        if swatch:
            ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor=swatch)
        row += 1
    # Fila de totales (suma YA calculada en el API; aquí solo se pinta).
    ws.cell(row=row, column=1, value="TOTALES").font = Font(bold=True)
    for col, val, fmt in (
        (4, req.combustible_litros, "0.0"),
        (5, req.combustible_total_mxn, MONEY),
    ):
        if val is not None:
            cell = ws.cell(row=row, column=col, value=round(val, 2))
            cell.font = Font(bold=True)
            cell.number_format = fmt
            cell.fill = PatternFill("solid", fgColor=LIGHT)
    if req.combustible_precio_litro is not None:
        cell = ws.cell(row=row, column=6, value=round(req.combustible_precio_litro, 2))
        cell.font = Font(bold=True)
        cell.number_format = MONEY
        cell.fill = PatternFill("solid", fgColor=LIGHT)
        ws.cell(row=row + 1, column=6, value="($ x litro prom)").font = Font(
            italic=True, size=8, color="5B6470"
        )
    fila_nota = row + 2
    nota = (
        "El combustible se controla POR AVIÓN y POR MES (fecha del gasto, con "
        "o sin vuelo) — mismo criterio que el reparto a socios."
    )
    if req.combustible_sin_avion:
        nota += (
            f" AVISO: {req.combustible_sin_avion} carga(s) SIN avión — "
            "asígnales aeronave en Combustibles (bloquean el pre-cierre)."
        )
    ws.cell(row=fila_nota, column=1, value=nota).font = Font(
        italic=True, size=9, color="5B6470"
    )
    ws.freeze_panes = "A4"


def _hoja_utilidades(ws: Worksheet, req: DineroXlsxRequest) -> None:
    ws.title = "utilidades"
    ws.cell(row=1, column=1, value="UTILIDADES").font = Font(bold=True, size=12, color=NAVY)
    headers = [
        "PERIODO", "CHARTERS", "OTROS\nINGRESOS", "UTILIDAD ANTES\nGASTOS",
        "OTROS\nGASTOS", "COMBUSTIBLE\nDEL MES",
    ]
    for a in req.utilidades_aviones:
        headers += [
            f"GASTOS\nINDIRECTOS\n{a.matricula}",
            f"OTROS GASTOS\n{a.matricula}",
            f"PERMISOS\n{a.matricula}",
            f"COMBUSTIBLE\n{a.matricula}",
        ]
    headers += ["UTILIDAD\nDESPUES GASTOS", "TIPO\nCAMBIO", "UTILIDAD\nUSD"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=8)
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border
        ws.column_dimensions[get_column_letter(i)].width = 14
    ws.row_dimensions[3].height = 34
    periodo = f"{_fecha(req.periodo_desde)} — {_fecha(req.periodo_hasta)}"
    vals: list[object] = [
        periodo,
        None,  # CHARTERS: depende del costo proveedor (regla pendiente)
        req.utilidades_otros_ingresos_mxn,
        None,  # UTILIDAD ANTES: depende de charters
        req.utilidades_otros_gastos_mxn,
        req.utilidades_combustible_mxn,
    ]
    for a in req.utilidades_aviones:
        vals += [a.gastos_indirectos_mxn, a.otros_gastos_mxn, a.permisos_mxn,
                 a.combustible_mxn]
    vals += [None, req.utilidades_tc, None]
    for i, v in enumerate(vals, start=1):
        cell = ws.cell(row=4, column=i, value=v)
        cell.border = _border
        if isinstance(v, (int, float)):
            cell.number_format = TC if i == len(vals) - 1 else MONEY
    ws.cell(
        row=6,
        column=1,
        value="CHARTERS y UTILIDAD dependen del costo por hora del proveedor y "
        "las comisiones (reglas pendientes de definir con el equipo): esas "
        "celdas van vacías a propósito — no se inventan números.",
    ).font = Font(italic=True, size=9, color="5B6470")
    ws.cell(
        row=7,
        column=1,
        value="COMBUSTIBLE DEL MES = suma de las columnas por avión + las "
        "cargas SIN avión (ver pestaña Combustible; bloquean el pre-cierre).",
    ).font = Font(italic=True, size=9, color="5B6470")
    ws.cell(
        row=8,
        column=1,
        value="OTROS GASTOS = total del mes; las columnas por avión muestran "
        "solo lo asignado/repartido a cada aeronave — la diferencia es gasto "
        "de la empresa VuelaTour (y partidas sin tipo de cambio).",
    ).font = Font(italic=True, size=9, color="5B6470")
    ws.cell(
        row=9,
        column=1,
        value="OTROS INGRESOS = ingreso de VuelaTour (TUAs/extras/pernocta + su "
        "IVA), no del avión (regla 28-ago-2026): ver hoja 'Otros ingresos'.",
    ).font = Font(italic=True, size=9, color="5B6470")


def render_dinero_xlsx(req: DineroXlsxRequest) -> bytes:
    wb = Workbook()
    _hoja_vuelos(wb.active, req)
    _hoja_otros_ingresos(wb.create_sheet(), req)
    _hoja_otros_gastos(wb.create_sheet(), req)
    _hoja_combustible(wb.create_sheet(), req)
    _hoja_utilidades(wb.create_sheet(), req)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
