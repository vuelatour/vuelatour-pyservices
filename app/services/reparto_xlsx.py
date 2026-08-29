"""Reporte mensual por avión en Excel (doc 5.10 / etapa 7).

Reutiliza el mismo payload del reparto de utilidades (RepartoPdfRequest) y arma
un libro con: una hoja resumen por avión con las 6 secciones (ingresos, gastos
directos/indirectos, permisos, otros, reserva overhaul, saldo) y el reparto por
socio de cada avión.
"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.schemas.reparto import (
    RepartoOtrosIngresosDesglose,
    RepartoPdfRequest,
    RepartoVueloLinea,
)

BRAND = "0F4C81"
LIGHT = "EEF2F7"
WHITE = "FFFFFF"
MONEY = '"$"#,##0.00'

_thin = Side(style="thin", color="D5DBE3")
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _pct(factor: float | None) -> str:
    """0.5 → '50 %' (hasta 2 decimales, sin ceros de más)."""
    return "—" if factor is None else f"{round(factor * 100, 2):g} %"


def _usd(value: float) -> str:
    return f"${value:,.2f}"


def _desglose_txt(d: RepartoOtrosIngresosDesglose | None) -> str | None:
    """'TUAs $100.00 · comisión del vendedor $1,000.00 · IVA $176.00' con la
    composición COTIZADA de otros ingresos VuelaTour; None si no hay nada."""
    if d is None:
        return None
    partes = [
        f"{label} {_usd(val)}"
        for label, val in (
            ("TUAs", d.tuas_usd),
            ("extras", d.extras_usd),
            ("pernocta", d.pernocta_usd),
            ("comisión del vendedor", d.comision_usd),
            ("IVA", d.iva_usd),
        )
        if val
    ]
    return " · ".join(partes) or None


def _nota_tc_oficial(avion) -> str | None:
    """Nota del TC oficial de respaldo del avión (29-ago-2026), o None."""
    partes = []
    g = getattr(avion, "gastos_tc_oficial", None)
    if g is not None and g.count > 0:
        partes.append(f"{g.count} gasto(s) MXN sin TC capturado (${g.monto_mxn:,.2f} MXN)")
    c = getattr(avion, "cobros_tc_oficial_count", 0) or 0
    if c > 0:
        partes.append(f"{c} vuelo(s) con cobros MXN sin TC")
    if not partes:
        return None
    return (
        " y ".join(partes)
        + " se convirtieron con el TC oficial de referencia del día "
        "(open.er-api / BCE); ya están dentro de las cifras."
    )


def _nota_tc_oficial_global(tc) -> str | None:
    """Nota del periodo (resumen): vuelos/gastos convertidos con TC oficial."""
    if tc is None:
        return None
    gastos = tc.gastos.count if tc.gastos is not None else 0
    if tc.vuelos <= 0 and gastos <= 0:
        return None
    partes = []
    if gastos > 0:
        partes.append(f"{gastos} gasto(s) MXN (${tc.gastos.monto_mxn:,.2f} MXN)")
    if tc.vuelos > 0:
        partes.append(f"{tc.vuelos} vuelo(s) con cobros MXN")
    fuentes = f" Fuente(s): {' / '.join(tc.fuentes)}." if tc.fuentes else ""
    return (
        "Tipo de cambio de referencia: "
        + " y ".join(partes)
        + " sin TC capturado se convirtieron con el TC oficial del día de la "
        "cotización o del gasto (open.er-api / BCE); ya están dentro de las cifras."
        + fuentes
    )


def _folio_vuelo(v: RepartoVueloLinea) -> str:
    """'#105 · 50 %' cuando el vuelo fue MULTI-AVIÓN (regla B, 28-ago-2026:
    la venta del avión se repartió por tramo); '#105' si no."""
    texto = f"#{v.folio}" if v.folio is not None else "—"
    if v.participacion is not None and v.participacion < 0.9999:
        texto += f" · {_pct(v.participacion)}"
    return texto


def _title(ws, text: str, row: int, span: int, size: int = 14):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(bold=True, size=size, color=BRAND)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)


def _header_row(ws, row: int, headers: list[str]):
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(bold=True, color=WHITE)
        c.fill = PatternFill("solid", fgColor=BRAND)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border


def render_reparto_xlsx(req: RepartoPdfRequest) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen por avión"

    n_cols = 14
    _title(ws, "VuelaTour — Reporte mensual por avión", 1, n_cols, size=16)
    sub = ws.cell(row=2, column=1, value=f"Periodo: {req.periodo_desde} a {req.periodo_hasta}  ·  Generado: {req.generado}")
    sub.font = Font(italic=True, size=10, color="5B6470")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)

    headers = [
        "Matrícula", "Modelo", "Horas voladas", "Venta del avión cobrada", "Comisiones venta",
        "Pendiente de cobro (parte avión)", "Gastos directos", "Gastos indirectos", "Permisos",
        "Fijos (prorrateo + reparto)", "Reserva overhaul", "Saldo disponible",
        # Informativas (regla 28-ago), FUERA de las columnas que suman el saldo:
        # TUAs/extras/pernocta cobrados = ingreso de VuelaTour; y la deuda
        # COMPLETA del cliente cuando es mayor a la parte del avión.
        "Otros ingr. VuelaTour (informativo)",
        "Deuda total del cliente (informativo)",
    ]
    hrow = 4
    _header_row(ws, hrow, headers)

    money_cols = list(range(4, 13))  # columnas D..L son montos (C = horas)
    totals = {c: 0.0 for c in money_cols}
    info_col = 13  # M: otros ingresos VuelaTour (total propio, no suma al saldo)
    total_info = 0.0
    # N: deuda total del cliente — solo cuando supera la parte del avión
    # (pendiente_bruto_usd > pendiente_cobro_usd); en gris, sin total.
    deuda_col = 14

    r = hrow + 1
    for a in req.aviones:
        valores = [
            a.matricula, a.modelo, a.horas_voladas_hr, a.ingresos_cobrado_usd,
            a.comisiones_venta_usd, a.pendiente_cobro_usd,
            a.gastos_directos_usd, a.gastos_indirectos_usd, a.permisos_usd,
            a.otros_usd, a.reserva_overhaul_usd, a.saldo_usd,
            a.otros_ingresos_vuelatour_usd,
        ]
        for col, v in enumerate(valores, start=1):
            c = ws.cell(row=r, column=col, value=v)
            c.border = _border
            if col == 3:
                c.number_format = "0.0"
            if col in money_cols:
                c.number_format = MONEY
                totals[col] += float(v or 0)
            if col == info_col:
                c.number_format = MONEY
                c.font = Font(italic=True, color="5B6470")
                total_info += float(v or 0)
        pendiente = a.pendiente_cobro_usd or 0.0
        bruto = a.pendiente_bruto_usd or 0.0
        dc = ws.cell(
            row=r, column=deuda_col,
            value=bruto if bruto > pendiente + 0.005 else None,
        )
        dc.number_format = MONEY
        dc.font = Font(italic=True, color="5B6470")
        dc.border = _border
        r += 1

    # Fila de totales.
    tcell = ws.cell(row=r, column=1, value="TOTAL")
    tcell.font = Font(bold=True)
    tcell.fill = PatternFill("solid", fgColor=LIGHT)
    ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor=LIGHT)
    for col in money_cols:
        c = ws.cell(row=r, column=col, value=totals[col])
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor=LIGHT)
        c.number_format = MONEY
        c.border = _border
    # Total propio de la columna informativa (no entra al saldo).
    ci = ws.cell(row=r, column=info_col, value=total_info)
    ci.font = Font(bold=True, italic=True, color="5B6470")
    ci.fill = PatternFill("solid", fgColor=LIGHT)
    ci.number_format = MONEY
    ci.border = _border
    cd = ws.cell(row=r, column=deuda_col)
    cd.fill = PatternFill("solid", fgColor=LIGHT)
    cd.border = _border
    r += 1
    # Regla A (28-ago-2026): la comisión del vendedor ya no se descuenta al
    # avión — el API manda 0. La columna se conserva por índice (money_cols,
    # widths) pero se OCULTA cuando nadie trae monto (solo un payload viejo
    # la muestra).
    if not any(a.comisiones_venta_usd for a in req.aviones):
        ws.column_dimensions[get_column_letter(5)].hidden = True
    nota = ws.cell(
        row=r,
        column=1,
        value="Venta del avión cobrada = tiempo de vuelo + ajuste + IVA proporcional "
        "(en vuelos multi-avión, la parte de cada matrícula en partes iguales por "
        "tramo vendido; los ferries/tramos operativos no reparten). "
        "Otros ingr. VuelaTour = TUAs/extras/pernocta/comisión del vendedor cobrados "
        "(con su IVA): ingreso de VuelaTour, fuera del saldo y del reparto; el pago "
        "de la comisión al vendedor sale de VuelaTour, no del avión (ver 'otros "
        "movimientos' del Balance general). Pendiente de cobro = parte del avión "
        "(se reparte al cobrar); Deuda total del cliente = lo que debe COMPLETO "
        "(con TUAs/extras/pernocta/comisión), solo cuando es mayor a la parte del "
        "avión.",
    )
    nota.font = Font(italic=True, size=9, color="5B6470")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_cols)
    # Composición cotizada de "Otros ingr. VuelaTour" (regla A: incluye la
    # comisión del vendedor) — global y por avión, solo si el API la manda.
    desglose_txt = _desglose_txt(req.otros_ingresos_vuelatour_desglose)
    if desglose_txt:
        r += 1
        nd = ws.cell(
            row=r,
            column=1,
            value="Otros ingr. VuelaTour del periodo (cotizado, no se reparte): "
            f"{desglose_txt}.",
        )
        nd.font = Font(italic=True, size=9, color="5B6470")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_cols)
    por_avion = [
        f"{a.matricula} {_usd(a.otros_ingresos_vuelatour_desglose.comision_usd)}"
        for a in req.aviones
        if a.otros_ingresos_vuelatour_desglose is not None
        and a.otros_ingresos_vuelatour_desglose.comision_usd
    ]
    if por_avion:
        r += 1
        nc = ws.cell(
            row=r,
            column=1,
            value="Comisión del vendedor cotizada por avión (ingreso de VuelaTour, "
            "su pago no es costo del avión): " + " · ".join(por_avion) + ".",
        )
        nc.font = Font(italic=True, size=9, color="5B6470")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_cols)
    # Nota global del TC oficial de respaldo (29-ago-2026): misma redacción
    # que el PDF.
    nota_global = _nota_tc_oficial_global(req.tc_oficial)
    if nota_global:
        r += 1
        nt = ws.cell(row=r, column=1, value=nota_global)
        nt.font = Font(italic=True, size=9, color="5B6470")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n_cols)

    # Anchos de columna.
    widths = [12, 16, 13, 16, 16, 17, 15, 16, 12, 16, 16, 16, 18, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Reparto por socio (debajo del resumen).
    r += 2
    _title(ws, "Reparto por socio", r, 4, size=12)
    r += 1
    _header_row(ws, r, ["Avión", "Socio", "Porcentaje", "Monto USD"])
    r += 1
    for a in req.aviones:
        for linea in a.reparto:
            ws.cell(row=r, column=1, value=a.matricula).border = _border
            ws.cell(row=r, column=2, value=linea.socio_nombre).border = _border
            pc = ws.cell(row=r, column=3, value=(linea.porcentaje or 0) / 100)
            pc.number_format = "0.00%"
            pc.border = _border
            mc = ws.cell(row=r, column=4, value=linea.monto_usd)
            mc.number_format = MONEY
            mc.border = _border
            r += 1
        # Vigencias traslapadas/incompletas: mismo aviso que el PDF y el badge
        # de la web — el Excel dispersable no puede salir doble en silencio.
        if a.reparto and round(a.reparto_porcentaje_total, 2) != 100:
            aviso = ws.cell(
                row=r,
                column=1,
                value=(
                    f"AVISO {a.matricula}: los porcentajes suman "
                    f"{a.reparto_porcentaje_total:g}% (no 100%) — revisar vigencias de socios."
                ),
            )
            aviso.font = Font(color="B45309", italic=True, size=9)
            r += 1
        # Nota informativa (29-ago-2026): montos MXN sin TC capturado que SÍ
        # entraron, convertidos con el TC oficial de referencia del día.
        nota_tc = _nota_tc_oficial(a)
        if nota_tc:
            nc = ws.cell(row=r, column=1, value=f"NOTA {a.matricula}: {nota_tc}")
            nc.font = Font(italic=True, size=9, color="5B6470")
            r += 1

    # Detalle de vuelos (solo si el API lo manda): el folio lleva ' · 50 %'
    # cuando el vuelo fue MULTI-AVIÓN (regla B, 28-ago-2026: la venta del
    # avión se repartió por tramo entre las matrículas).
    if any(a.vuelos for a in req.aviones):
        r += 2
        _title(ws, "Detalle de vuelos (venta del avión cobrada)", r, 6, size=12)
        r += 1
        _header_row(ws, r, ["Avión", "Vuelo", "Cliente", "Cobrado avión USD",
                            "Pendiente avión USD", "Tramos del avión"])
        r += 1
        for a in req.aviones:
            for v in a.vuelos:
                ws.cell(row=r, column=1, value=a.matricula).border = _border
                ws.cell(row=r, column=2, value=_folio_vuelo(v)).border = _border
                ws.cell(row=r, column=3, value=v.cliente).border = _border
                for col, val in ((4, v.cobrado_usd), (5, v.pendiente_usd)):
                    mc = ws.cell(row=r, column=col, value=val)
                    mc.number_format = MONEY
                    mc.border = _border
                ws.cell(row=r, column=6, value=v.tramos_avion).border = _border
                r += 1

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
