"""Generador genérico de Excel a partir de una tabla (titulo + columnas + filas).

Lo usan los exports del sistema (inventario valorizado, cardex, horas por
piloto, gastos por avión/categoría, etc.) para no duplicar formato.
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.schemas.tabla import TablaXlsxRequest

# Excel prohíbe estos caracteres en el NOMBRE de la hoja (no en las celdas):
# un titulo como "Gastos por avión / categoría" tronaba el export completo.
_SHEET_ILEGALES = str.maketrans({c: "-" for c in "\\/*?:[]"})


def sheet_title(nombre: str, respaldo: str = "Reporte") -> str:
    """Nombre de hoja válido para Excel (sin caracteres ilegales, máx 31)."""
    limpio = (nombre or respaldo).translate(_SHEET_ILEGALES).strip()
    return (limpio or respaldo)[:31]

BRAND = "0F4C81"
LIGHT = "EEF2F7"
WHITE = "FFFFFF"
MONEY = '"$"#,##0.00'

_thin = Side(style="thin", color="D5DBE3")
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _fmt_for(tipo: str) -> str | None:
    return {
        "money": MONEY,
        "numero": "#,##0.00",
        "entero": "#,##0",
        "pct": "0.00%",
    }.get(tipo)


def render_tabla_xlsx(req: TablaXlsxRequest) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title(req.titulo)

    ncols = max(len(req.columnas), 1)

    # Título.
    t = ws.cell(row=1, column=1, value=req.titulo)
    t.font = Font(bold=True, size=14, color=BRAND)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    hrow = 2
    if req.subtitulo:
        s = ws.cell(row=2, column=1, value=req.subtitulo)
        s.font = Font(italic=True, size=10, color="5B6470")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        hrow = 3

    # Bloque RESUMEN (opcional) arriba de la tabla: pares etiqueta → valor,
    # p. ej. el total del periodo por categoría de gasto.
    if req.resumen:
        hrow += 1
        if req.resumen_titulo:
            rt = ws.cell(row=hrow, column=1, value=req.resumen_titulo)
            rt.font = Font(bold=True, size=11, color=BRAND)
            hrow += 1
        for par in req.resumen:
            etiqueta = par[0] if len(par) > 0 else None
            valor = par[1] if len(par) > 1 else None
            e = ws.cell(row=hrow, column=1, value=etiqueta)
            e.fill = PatternFill("solid", fgColor=LIGHT)
            e.border = _border
            v = ws.cell(row=hrow, column=2, value=valor)
            v.border = _border
            if isinstance(valor, (int, float)):
                v.number_format = MONEY
            # Celdas extra del par (p. ej. conteo) tal cual.
            for col in range(3, len(par) + 1):
                x = ws.cell(row=hrow, column=col, value=par[col - 1])
                x.border = _border
            hrow += 1

    hrow += 1  # una fila en blanco antes del encabezado

    # Encabezado.
    for col, c in enumerate(req.columnas, start=1):
        cell = ws.cell(row=hrow, column=col, value=c.label)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=BRAND)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border
        ws.column_dimensions[get_column_letter(col)].width = max(12, len(c.label) + 4)

    # Filas.
    r = hrow + 1
    for fila in req.filas:
        for col, (c, v) in enumerate(zip(req.columnas, fila), start=1):
            cell = ws.cell(row=r, column=col, value=v)
            cell.border = _border
            fmt = _fmt_for(c.tipo)
            if fmt and v is not None:
                cell.number_format = fmt
        r += 1

    # Totales (opcional).
    if req.totales:
        for col, (c, v) in enumerate(zip(req.columnas, req.totales), start=1):
            cell = ws.cell(row=r, column=col, value=v)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor=LIGHT)
            cell.border = _border
            fmt = _fmt_for(c.tipo)
            if fmt and v is not None:
                cell.number_format = fmt

    ws.freeze_panes = ws.cell(row=hrow + 1, column=1)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
