"""Cardex de un ítem en formato LIBRO (réplica del cuaderno del cliente).

Un solo sheet con dos bloques lado a lado — ENTRADAS | SALIDAS — cada uno
con sus columnas (venta, remanente y ganancia del lado de salidas). Todo
viene YA calculado del API: aquí solo se pinta el libro con los estilos
BRAND de la casa (tabla_xlsx).
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.schemas.cardex_libro import CardexLibroRequest
from app.services.tabla_xlsx import BRAND, LIGHT, MONEY, WHITE, sheet_title

_thin = Side(style="thin", color="D5DBE3")
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
# Separación visual entre el bloque ENTRADAS y el de SALIDAS.
_sep = Side(style="medium", color=BRAND)

# Columnas del bloque ENTRADAS (1-6) y del bloque SALIDAS (7-14).
_COLS_ENTRADAS = [
    ("Fecha entrada", 12),
    ("Cantidad", 10),
    ("Descripción", 34),
    ("Valor compra", 13),
    ("Valor compra total", 14),
    ("Cantidad en stock", 11),
]
_COLS_SALIDAS = [
    ("Fecha de salida", 12),
    ("Cantidad", 10),
    ("Descripción del producto", 34),
    ("Valor unitario al que se vendió", 14),
    ("Valor total", 13),
    ("Remanente", 11),
    ("Ganancia", 13),
    ("A quién se le vendió", 16),
]
_N_ENT = len(_COLS_ENTRADAS)  # 6
_NCOLS = _N_ENT + len(_COLS_SALIDAS)  # 14
# Columnas con formato de dinero (1-based).
_MONEY_COLS = {4, 5, _N_ENT + 4, _N_ENT + 5, _N_ENT + 7}


def _celda(ws, row: int, col: int, value, *, money: bool = False, wrap: bool = False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.border = _border
    if money and value is not None:
        cell.number_format = MONEY
    if wrap:
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    # Borde grueso entre bloques (última col de ENTRADAS | primera de SALIDAS).
    if col == _N_ENT:
        cell.border = Border(left=_thin, right=_sep, top=_thin, bottom=_thin)
    return cell


def render_cardex_libro_xlsx(req: CardexLibroRequest) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title("Cardex libro")

    # Título + subtítulo.
    t = ws.cell(row=1, column=1, value=req.titulo or f"Cardex — {req.item_nombre}")
    t.font = Font(bold=True, size=14, color=BRAND)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=_NCOLS)
    partes = []
    if req.numero_parte:
        partes.append(f"Parte {req.numero_parte}")
    if req.unidad:
        partes.append(f"Unidad: {req.unidad}")
    partes.append(f"Montos en {req.moneda or 'MXN'}")
    if req.generado:
        partes.append(f"Generado {req.generado}")
    s = ws.cell(row=2, column=1, value=" · ".join(partes))
    s.font = Font(italic=True, size=10, color="5B6470")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=_NCOLS)

    # Encabezado DOBLE: fila de bloques (ENTRADAS | SALIDAS) + fila de columnas.
    hrow = 4
    for etiqueta, c1, c2 in (
        ("ENTRADAS", 1, _N_ENT),
        ("SALIDAS", _N_ENT + 1, _NCOLS),
    ):
        cell = ws.cell(row=hrow, column=c1, value=etiqueta)
        cell.font = Font(bold=True, color=WHITE, size=12)
        cell.fill = PatternFill("solid", fgColor=BRAND)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        for col in range(c1, c2 + 1):
            x = ws.cell(row=hrow, column=col)
            x.fill = PatternFill("solid", fgColor=BRAND)
            x.border = _border
        ws.merge_cells(start_row=hrow, start_column=c1, end_row=hrow, end_column=c2)

    crow = hrow + 1
    for col, (label, width) in enumerate(_COLS_ENTRADAS + _COLS_SALIDAS, start=1):
        cell = ws.cell(row=crow, column=col, value=label)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=BRAND)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border
        ws.column_dimensions[get_column_letter(col)].width = width

    # Filas: cada lado avanza por su cuenta (como en el cuaderno físico).
    r = crow + 1
    nfilas = max(len(req.entradas), len(req.salidas))
    for i in range(nfilas):
        e = req.entradas[i] if i < len(req.entradas) else None
        v = req.salidas[i] if i < len(req.salidas) else None
        _celda(ws, r, 1, e.fecha if e else None)
        _celda(ws, r, 2, e.cantidad if e else None)
        _celda(ws, r, 3, e.descripcion if e else None, wrap=True)
        _celda(ws, r, 4, e.valor_compra_unitario if e else None, money=True)
        _celda(ws, r, 5, e.valor_compra_total if e else None, money=True)
        _celda(ws, r, 6, e.stock_despues if e else None)
        _celda(ws, r, 7, v.fecha if v else None)
        _celda(ws, r, 8, v.cantidad if v else None)
        _celda(ws, r, 9, v.descripcion if v else None, wrap=True)
        _celda(ws, r, 10, v.venta_unitaria if v else None, money=True)
        _celda(ws, r, 11, v.venta_total if v else None, money=True)
        _celda(ws, r, 12, v.remanente if v else None)
        _celda(ws, r, 13, v.ganancia if v else None, money=True)
        _celda(ws, r, 14, v.vendido_a if v else None)
        r += 1

    # Totales (los manda el API): compra | venta | ganancia.
    tot = {
        3: "TOTAL COMPRA",
        5: req.total_compra,
        9: "TOTAL VENTA / GANANCIA",
        11: req.total_venta,
        13: req.total_ganancia,
    }
    for col in range(1, _NCOLS + 1):
        cell = ws.cell(row=r, column=col, value=tot.get(col))
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor=LIGHT)
        cell.border = _border
        if col in _MONEY_COLS and tot.get(col) is not None:
            cell.number_format = MONEY

    ws.freeze_panes = ws.cell(row=crow + 1, column=1)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
