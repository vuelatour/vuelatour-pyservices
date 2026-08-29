"""Alta masiva de inventario: plantilla XLSX y parseo estructural.

La oficina descarga la plantilla (una fila = un ítem de bodega con su
empaque/caja opcional y su existencia inicial), la llena y la sube. Aquí SOLO
se renderiza el libro y se convierte de vuelta a filas crudas con tipos
básicos; toda la validación de negocio (nombre/categoría obligatorios,
códigos únicos, costo/TC de la entrada inicial) vive en vuelatour-api.

Calcado de combustible_masivo.py: ENCABEZADOS = contrato de la fila 1, hoja
"Catálogos" con dropdowns, fila de ejemplo en gris y parser por NOMBRE de
encabezado (tolerante a acentos, mayúsculas, asteriscos y paréntesis).
"""

from __future__ import annotations

import base64
import binascii
import csv
import io
import unicodedata
from collections.abc import Callable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from app.schemas.inventario import (
    FilaInventario,
    ParseInventarioRequest,
    ParseInventarioResponse,
    PlantillaInventarioRequest,
)

HOJA_ITEMS = "Inventario"
HOJA_CATALOGOS = "Catálogos"
# Las validaciones de datos y el formato texto cubren las filas 2..500.
MAX_FILA_DV = 500

BRAND = "0F4C81"
GRIS_EJEMPLO = "9AA0A6"
WHITE = "FFFFFF"

# Contrato de columnas: ORDEN EXACTO de la fila 1 (el API y el panel dependen
# de estos encabezados; el parser los busca por nombre, no por posición).
ENCABEZADOS = [
    "Nombre*",
    "Marca",
    "Categoría*",
    "Número de parte",
    "Código de barras (unidad)",
    "Unidad de medida",
    "Descripción",
    "Ubicación",
    "Stock mínimo",
    "Existencia inicial",
    "Costo unitario",
    "Moneda",
    "Tipo de cambio",
    "Empaque (nombre)",
    "Unidades por empaque",
    "Código de barras del empaque",
    "Notas",
]
# Campo de FilaInventario que corresponde a cada encabezado (mismo orden).
CAMPOS = [
    "nombre",
    "marca",
    "categoria",
    "numero_parte",
    "codigo",
    "unidad",
    "descripcion",
    "ubicacion",
    "stock_minimo",
    "existencia_inicial",
    "costo_unitario",
    "moneda",
    "tipo_cambio",
    "empaque_nombre",
    "empaque_factor",
    "empaque_codigo",
    "notas",
]
ANCHOS = [30, 14, 16, 16, 24, 15, 46, 16, 12, 14, 14, 10, 14, 16, 12, 26, 32]

# Columnas (1-based) que deben quedar como TEXTO ('@'): Excel convertiría
# los códigos en número o notación científica y perdería ceros a la izquierda.
COLS_TEXTO = (4, 5, 16)  # Número de parte, Código unidad, Código empaque
COLS_MONTO = (11, 13)  # Costo unitario, Tipo de cambio

# Catálogos de respaldo cuando el API no manda lista (unidad y moneda son
# estables; las categorías son del negocio y sin catálogo quedan libres).
UNIDADES_DEFAULT = ["pieza", "botella", "litro", "galón", "juego", "kit", "caja", "metro"]
MONEDAS_DEFAULT = ["MXN", "USD"]

# Fila de ejemplo (gris, itálica). Es un producto REAL de la bodega: el
# parser la descarta SOLO si viene intacta (todas las celdas iguales); si el
# usuario la edita se toma como fila válida.
NOTA_EJEMPLO = "(borra esta fila de ejemplo)"
EJEMPLO: list[object] = [
    "AeroShell W15W-50",
    "AeroShell",
    "Aceites",
    "550050835",
    "021400062153",
    "botella",
    "Aceite semisintético para motores de pistón, 1 qt (946 mL)",
    "Bodega Cancún",
    6,
    12,
    350,
    "MXN",
    17.5,
    "Caja de 6",
    6,
    "00021400062160",
    NOTA_EJEMPLO,
]

INSTRUCCIONES = [
    "Una fila = un producto/refacción de bodega (ítem).",
    "Los campos con * son obligatorios: Nombre y Categoría.",
    "Código de barras (unidad): los dígitos impresos bajo las barras del producto, sin "
    "espacios (ej. 021400062153). La columna está en formato TEXTO para conservar los "
    "ceros a la izquierda: no le cambies el formato.",
    "Número de parte: el código del fabricante impreso en la etiqueta o caja "
    "(ej. Product code 550050835); no es el código de barras.",
    "Empaque: si el producto también se maneja por caja, captura el nombre del empaque "
    "(ej. Caja de 6), cuántas unidades trae y el código de barras propio de la caja "
    "(ej. 00021400062160). Déjalo vacío si solo se maneja por unidad.",
    "Existencia inicial: unidades que hay hoy en bodega; con ella se genera la ENTRADA "
    "inicial del cardex usando Costo unitario, Moneda y Tipo de cambio (MXN por USD).",
    "Stock mínimo: al bajar de esa cantidad el sistema avisa para reabastecer.",
    "Categoría, Unidad de medida y Moneda tienen lista desplegable (hoja Catálogos). "
    "Si necesitas una categoría o unidad nueva, escríbela y acepta el aviso.",
    "Borra la fila de ejemplo (en gris) antes de subir el archivo.",
    "No cambies los encabezados ni el orden de las columnas.",
]


# --------------------------------------------------------------------------
# Plantilla
# --------------------------------------------------------------------------


def _catalogo(valores: list[str], default: list[str]) -> list[str]:
    limpios = [v.strip() for v in valores if v and v.strip()]
    return limpios or list(default)


def render_plantilla_inventario(req: PlantillaInventarioRequest) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = HOJA_ITEMS

    # Encabezado (fila 1).
    for col, (titulo, ancho) in enumerate(zip(ENCABEZADOS, ANCHOS, strict=True), start=1):
        cell = ws.cell(row=1, column=col, value=titulo)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=BRAND)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = ancho
    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

    # Formato TEXTO en las columnas de códigos (filas 2..MAX_FILA_DV) ANTES de
    # escribir el ejemplo: así lo que teclee el usuario se conserva tal cual.
    for col in COLS_TEXTO:
        for r in range(2, MAX_FILA_DV + 1):
            ws.cell(row=r, column=col).number_format = "@"

    # Fila 2: ejemplo realista (gris itálica) que el usuario debe borrar.
    fuente_ejemplo = Font(italic=True, color=GRIS_EJEMPLO)
    for col, valor in enumerate(EJEMPLO, start=1):
        cell = ws.cell(row=2, column=col, value=valor)
        cell.font = fuente_ejemplo
        if col in COLS_MONTO:
            cell.number_format = "#,##0.00"

    # Hoja de catálogos: de aquí salen los dropdowns (rangos absolutos).
    categorias = _catalogo(req.categorias, [])
    unidades = _catalogo(req.unidades, UNIDADES_DEFAULT)
    monedas = _catalogo(req.monedas, MONEDAS_DEFAULT)
    cat = wb.create_sheet(HOJA_CATALOGOS)
    catalogos = [
        ("CATEGORÍAS", categorias),
        ("UNIDADES DE MEDIDA", unidades),
        ("MONEDAS", monedas),
    ]
    for col, (titulo, valores) in enumerate(catalogos, start=1):
        head = cat.cell(row=1, column=col, value=titulo)
        head.font = Font(bold=True, color=WHITE)
        head.fill = PatternFill("solid", fgColor=BRAND)
        ancho = max([len(titulo)] + [len(v) for v in valores]) + 4
        cat.column_dimensions[get_column_letter(col)].width = ancho
        for r, valor in enumerate(valores, start=2):
            cat.cell(row=r, column=col, value=valor)
    cat.freeze_panes = "A2"

    # Validaciones de datos (dropdowns) sobre las filas 2..MAX_FILA_DV.
    # columna en "Inventario" -> (columna del catálogo, valores, estilo de error).
    # Categoría y unidad admiten valores nuevos (aviso, no bloqueo): el API
    # decide si crea la categoría. Moneda sí es cerrada.
    dropdowns = [
        ("C", 1, len(categorias), "warning"),  # Categoría
        ("F", 2, len(unidades), "warning"),  # Unidad de medida
        ("L", 3, len(monedas), "stop"),  # Moneda
    ]
    for col_item, col_cat, n, estilo in dropdowns:
        if n == 0:
            continue  # sin catálogo no hay dropdown (campo queda libre)
        letra_cat = get_column_letter(col_cat)
        dv = DataValidation(
            type="list",
            formula1=f"'{HOJA_CATALOGOS}'!${letra_cat}$2:${letra_cat}${n + 1}",
            allow_blank=True,
            showErrorMessage=True,
            errorStyle=estilo,
        )
        if estilo == "stop":
            dv.errorTitle = "Valor inválido"
            dv.error = "Elige un valor de la lista (hoja Catálogos)."
        else:
            dv.errorTitle = "Valor fuera del catálogo"
            dv.error = (
                "Ese valor no está en la hoja Catálogos. ¿Quieres usarlo de todos modos "
                "(se creará como nuevo)?"
            )
        ws.add_data_validation(dv)
        dv.add(f"{col_item}2:{col_item}{MAX_FILA_DV}")

    # Hoja de instrucciones.
    ins = wb.create_sheet("Instrucciones")
    titulo = ins.cell(row=1, column=1, value="INSTRUCCIONES — ALTA MASIVA DE INVENTARIO")
    titulo.font = Font(bold=True, size=13, color=BRAND)
    ins.column_dimensions["A"].width = 110
    for r, linea in enumerate(INSTRUCCIONES, start=3):
        c = ins.cell(row=r, column=1, value=f"• {linea}")
        c.alignment = Alignment(wrap_text=True, vertical="top")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --------------------------------------------------------------------------
# Parseo (estructural, sin validación de negocio)
# --------------------------------------------------------------------------


def _normaliza(texto: str) -> str:
    """minúsculas, sin acentos, sin '*', sin '(...)', espacios colapsados."""
    s = unicodedata.normalize("NFD", texto)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = s.lower().replace("*", " ")
    while "(" in s:
        a = s.index("(")
        b = s.find(")", a)
        s = s[:a] + (" " + s[b + 1 :] if b != -1 else " ")
    return " ".join(s.split())


# Encabezado normalizado -> campo. Tolera variantes razonables.
_ALIAS = {
    "nombre": "nombre",
    "producto": "nombre",
    "nombre del producto": "nombre",
    "articulo": "nombre",
    "item": "nombre",
    "marca": "marca",
    "fabricante": "marca",
    "categoria": "categoria",
    "numero de parte": "numero_parte",
    "numero parte": "numero_parte",
    "no. de parte": "numero_parte",
    "no de parte": "numero_parte",
    "part number": "numero_parte",
    "product code": "numero_parte",
    "codigo de fabricante": "numero_parte",
    "codigo de barras": "codigo",  # "Código de barras (unidad)" sin paréntesis
    "codigo de barras unidad": "codigo",
    "codigo": "codigo",
    "codigo unidad": "codigo",
    "sku": "codigo",
    "upc": "codigo",
    "ean": "codigo",
    "unidad de medida": "unidad",
    "unidad": "unidad",
    "um": "unidad",
    "descripcion": "descripcion",
    "ubicacion": "ubicacion",
    "stock minimo": "stock_minimo",
    "stock min": "stock_minimo",
    "minimo": "stock_minimo",
    "existencia inicial": "existencia_inicial",
    "existencia": "existencia_inicial",
    "stock inicial": "existencia_inicial",
    "cantidad": "existencia_inicial",
    "costo unitario": "costo_unitario",
    "costo": "costo_unitario",
    "precio unitario": "costo_unitario",
    "moneda": "moneda",
    "tipo de cambio": "tipo_cambio",
    "tc": "tipo_cambio",
    "empaque": "empaque_nombre",  # "Empaque (nombre)" sin paréntesis
    "empaque nombre": "empaque_nombre",
    "nombre del empaque": "empaque_nombre",
    "unidades por empaque": "empaque_factor",
    "piezas por empaque": "empaque_factor",
    "unidades por caja": "empaque_factor",
    "factor": "empaque_factor",
    "codigo de barras del empaque": "empaque_codigo",
    "codigo de barras empaque": "empaque_codigo",
    "codigo del empaque": "empaque_codigo",
    "codigo empaque": "empaque_codigo",
    "notas": "notas",
    "nota": "notas",
    "observaciones": "notas",
}

_AVISO_CODIGO_NUMERICO = (
    "venía como NÚMERO en la celda (pudo perder ceros a la izquierda); verifícalo."
)


def _texto(v: object) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _numero(v: object) -> float | str | None:
    if v is None:
        return None
    if isinstance(v, bool):
        return _texto(v)
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    limpio = s.replace("$", "").replace(",", "").replace(" ", "")
    try:
        return float(limpio)
    except ValueError:
        return s  # crudo: el API lo reporta como ilegible


def _entero_texto(v: object) -> str | None:
    """Texto; si la celda era numérica entera, el entero sin '.0' (número de parte)."""
    if isinstance(v, bool):
        return _texto(v)
    if isinstance(v, float) and v.is_integer() and abs(v) < 1e15:
        return str(int(v))
    if isinstance(v, int):
        return str(v)
    return _texto(v)


def _codigo(v: object) -> tuple[str | None, str | None]:
    """Código de barras SIEMPRE como str sin espacios; (valor, aviso).

    - Texto: se quitan TODOS los espacios ('0 21400 06215 3' -> '021400062153').
      Una cadena en notación científica ('2.14E+10') se devuelve cruda para que
      el API la reporte como ilegible.
    - Número (el usuario tecleó los dígitos en una celda sin formato texto):
      se reconstruye el entero sin '.0' y se avisa que pudo perder ceros.
    """
    if v is None:
        return None, None
    if isinstance(v, bool):
        return _texto(v), None
    if isinstance(v, int):
        return str(v), _AVISO_CODIGO_NUMERICO
    if isinstance(v, float):
        if v.is_integer() and abs(v) < 1e15:
            return str(int(v)), _AVISO_CODIGO_NUMERICO
        return str(v), _AVISO_CODIGO_NUMERICO
    s = "".join(str(v).split())
    return (s or None), None


def _leer_xlsx(raw: bytes) -> tuple[list, list[tuple[int, list]]]:
    try:
        wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    except Exception as e:  # zip corrupto, formato viejo, etc.
        raise ValueError(
            "No se pudo leer el archivo Excel. Sube el .xlsx de la plantilla "
            "de inventario (o un .csv)."
        ) from e
    try:
        objetivo = _normaliza(HOJA_ITEMS)
        nombre = next(
            (n for n in wb.sheetnames if _normaliza(n) == objetivo), wb.sheetnames[0]
        )
        ws = wb[nombre]
        encabezado: list | None = None
        filas: list[tuple[int, list]] = []
        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if encabezado is None:
                if any(c is not None and str(c).strip() for c in row):
                    encabezado = list(row)
                continue
            filas.append((idx, list(row)))
    finally:
        wb.close()
    if encabezado is None:
        raise ValueError("El archivo está vacío: no se encontró la fila de encabezados.")
    return encabezado, filas


def _leer_csv(raw: bytes) -> tuple[list, list[tuple[int, list]]]:
    try:
        texto = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        texto = raw.decode("latin-1")
    try:
        dialecto: type[csv.Dialect] | csv.Dialect = csv.Sniffer().sniff(
            texto[:4096], delimiters=",;\t"
        )
    except csv.Error:
        dialecto = csv.excel
    encabezado: list | None = None
    filas: list[tuple[int, list]] = []
    for idx, row in enumerate(csv.reader(io.StringIO(texto), dialecto), start=1):
        if encabezado is None:
            if any(str(c).strip() for c in row):
                encabezado = list(row)
            continue
        filas.append((idx, list(row)))
    if encabezado is None:
        raise ValueError("El CSV está vacío: no se encontró la fila de encabezados.")
    return encabezado, filas


def _construye_fila(num: int, val: Callable[[str], object]) -> FilaInventario:
    """Convierte los valores crudos (por campo) en FilaInventario con avisos."""
    avisos: list[str] = []
    codigo, aviso = _codigo(val("codigo"))
    if aviso:
        avisos.append(f"Código de barras (unidad) {aviso}")
    empaque_codigo, aviso = _codigo(val("empaque_codigo"))
    if aviso:
        avisos.append(f"Código de barras del empaque {aviso}")
    return FilaInventario(
        fila=num,
        nombre=_texto(val("nombre")),
        marca=_texto(val("marca")),
        categoria=_texto(val("categoria")),
        numero_parte=_entero_texto(val("numero_parte")),
        codigo=codigo,
        unidad=_texto(val("unidad")),
        descripcion=_texto(val("descripcion")),
        ubicacion=_texto(val("ubicacion")),
        stock_minimo=_numero(val("stock_minimo")),
        existencia_inicial=_numero(val("existencia_inicial")),
        costo_unitario=_numero(val("costo_unitario")),
        moneda=_texto(val("moneda")),
        tipo_cambio=_numero(val("tipo_cambio")),
        empaque_nombre=_texto(val("empaque_nombre")),
        empaque_factor=_numero(val("empaque_factor")),
        empaque_codigo=empaque_codigo,
        notas=_texto(val("notas")),
        avisos=avisos,
    )


_EJEMPLO_POR_CAMPO: dict[str, object] = dict(zip(CAMPOS, EJEMPLO, strict=True))
_CAMPOS_COMPARABLES = {"fila", "avisos"}
# La fila de ejemplo pasada por el MISMO parser: así la comparación es exacta
# aunque los numéricos vuelvan como float (6 -> 6.0) o el CSV los traiga como texto.
_FILA_EJEMPLO = _construye_fila(0, _EJEMPLO_POR_CAMPO.get).model_dump(
    exclude=_CAMPOS_COMPARABLES
)


def _es_ejemplo_intacto(fila: FilaInventario) -> bool:
    return fila.model_dump(exclude=_CAMPOS_COMPARABLES) == _FILA_EJEMPLO


def parse_inventario(req: ParseInventarioRequest) -> ParseInventarioResponse:
    try:
        raw = base64.b64decode(req.archivo_base64, validate=True)
    except (binascii.Error, ValueError) as e:
        raise ValueError("El archivo no llegó completo (base64 inválido).") from e
    if not raw:
        raise ValueError("El archivo está vacío.")

    if (req.filename or "").lower().endswith(".csv"):
        encabezado, filas = _leer_csv(raw)
    else:
        encabezado, filas = _leer_xlsx(raw)

    # Mapa campo -> índice de columna, por ENCABEZADO (no por posición).
    indices: dict[str, int] = {}
    for i, celda in enumerate(encabezado):
        if celda is None:
            continue
        campo = _ALIAS.get(_normaliza(str(celda)))
        if campo and campo not in indices:
            indices[campo] = i
    if "nombre" not in indices and "codigo" not in indices:
        raise ValueError(
            "No se reconocieron los encabezados. ¿Es la plantilla de "
            "'Inventario'? No cambies los títulos de las columnas."
        )

    resultado: list[FilaInventario] = []
    for num, valores in filas:
        if not any(v is not None and str(v).strip() for v in valores):
            continue  # fila totalmente vacía

        def _val(campo: str, valores: list = valores) -> object:
            i = indices.get(campo)
            if i is None or i >= len(valores):
                return None
            return valores[i]

        fila = _construye_fila(num, _val)
        if fila.nombre is None and fila.codigo is None and fila.numero_parte is None:
            continue  # nada identifica un ítem (ni nombre, ni código, ni no. de parte)
        if _es_ejemplo_intacto(fila):
            continue  # el usuario no borró la fila de ejemplo: no es un alta
        resultado.append(fila)
    return ParseInventarioResponse(filas=resultado)
